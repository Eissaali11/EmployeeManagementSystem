"""
وحدة لوحة معلومات الحضور المحسنة
-----------------------------
تعرض إحصائيات الحضور حسب القسم والتاريخ مع إمكانية تصدير البيانات
"""

from datetime import datetime, timedelta
import pandas as pd
from sqlalchemy import func
from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required, current_user
import io
import tempfile
import os
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment

from models import Department, Employee, Attendance, Module
from app import db
from utils.decorators import module_access_required

# إنشاء blueprint للوحة معلومات الحضور
attendance_dashboard_bp = Blueprint('attendance_dashboard', __name__)

@attendance_dashboard_bp.route('/')
@login_required
@module_access_required(Module.ATTENDANCE)
def index():
    """صفحة لوحة معلومات الحضور الرئيسية"""
    # التحقق من التاريخ المحدد أو استخدام اليوم الحالي
    date_str = request.args.get('date')
    department_id = request.args.get('department_id')
    
    try:
        if date_str:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            selected_date = datetime.now().date()
    except ValueError:
        selected_date = datetime.now().date()
    
    # الحصول على الأقسام المتاحة للمستخدم حسب صلاحياته
    accessible_dept_ids = current_user.get_accessible_department_ids()
    
    if accessible_dept_ids is None:  # مدير النظام - جميع الأقسام
        departments = Department.query.order_by(Department.name).all()
    elif accessible_dept_ids:  # أقسام محددة
        departments = Department.query.filter(Department.id.in_(accessible_dept_ids)).order_by(Department.name).all()
    else:  # لا توجد أقسام متاحة
        departments = []
    
    # إحصائيات إجمالية - عدد الموظفين النشطين حسب الأقسام المتاحة
    if accessible_dept_ids is None:  # مدير النظام - جميع الموظفين
        total_employees = Employee.query.filter_by(status='active').count()
        
        # حساب إحصائيات الحضور الإجمالية
        total_present = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'present'
        ).scalar() or 0
        
        total_absent = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'absent'
        ).scalar() or 0
        
        total_leave = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'leave'
        ).scalar() or 0
        
        total_sick = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'sick'
        ).scalar() or 0
        
    elif accessible_dept_ids:  # أقسام محددة
        total_employees = Employee.query.filter(
            Employee.status == 'active',
            Employee.department_id.in_(accessible_dept_ids)
        ).count()
        
        # حساب إحصائيات الحضور الإجمالية للأقسام المتاحة
        total_present = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.status == 'active',
            Employee.department_id.in_(accessible_dept_ids),
            Attendance.date == selected_date,
            Attendance.status == 'present'
        ).scalar() or 0
        
        total_absent = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.status == 'active',
            Employee.department_id.in_(accessible_dept_ids),
            Attendance.date == selected_date,
            Attendance.status == 'absent'
        ).scalar() or 0
        
        total_leave = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.status == 'active',
            Employee.department_id.in_(accessible_dept_ids),
            Attendance.date == selected_date,
            Attendance.status == 'leave'
        ).scalar() or 0
        
        total_sick = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.status == 'active',
            Employee.department_id.in_(accessible_dept_ids),
            Attendance.date == selected_date,
            Attendance.status == 'sick'
        ).scalar() or 0
        
    else:  # لا توجد أقسام متاحة
        total_employees = 0
        total_present = 0
        total_absent = 0
        total_leave = 0
        total_sick = 0
    
    # حساب عدد الموظفين الذين لم يتم تسجيل حضورهم
    total_registered = total_present + total_absent + total_leave + total_sick
    total_missing = total_employees - total_registered if total_registered <= total_employees else 0
    
    # إحصائيات حسب القسم
    department_stats = []
    
    # جمع إحصائيات كل قسم
    for dept in departments:
        # عدد الموظفين في القسم
        employees_count = Employee.query.filter_by(
            department_id=dept.id, 
            status='active'
        ).count()
        
        # عدد الحاضرين
        present_count = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.department_id == dept.id,
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'present'
        ).scalar() or 0
        
        # عدد الغائبين
        absent_count = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.department_id == dept.id,
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'absent'
        ).scalar() or 0
        
        # عدد الإجازات
        leave_count = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.department_id == dept.id,
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'leave'
        ).scalar() or 0
        
        # عدد المرضي
        sick_count = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.department_id == dept.id,
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'sick'
        ).scalar() or 0
        
        # إجمالي السجلات
        total_records = present_count + absent_count + leave_count + sick_count
        
        # حساب نسب الحضور والغياب
        if employees_count > 0:
            present_percentage = round((present_count / employees_count) * 100, 1)
            absent_percentage = round((absent_count / employees_count) * 100, 1)
            missing_count = employees_count - total_records
            missing_percentage = round((missing_count / employees_count) * 100, 1) if missing_count > 0 else 0
        else:
            present_percentage = 0
            absent_percentage = 0
            missing_count = 0
            missing_percentage = 0
        
        # إضافة إحصائيات القسم للقائمة
        department_stats.append({
            'id': dept.id,
            'name': dept.name,
            'employees_count': employees_count,
            'present_count': present_count,
            'absent_count': absent_count,
            'leave_count': leave_count,
            'sick_count': sick_count,
            'missing_count': missing_count,
            'present_percentage': present_percentage,
            'absent_percentage': absent_percentage,
            'missing_percentage': missing_percentage
        })
    
    # الموظفون الغائبون
    absent_employees = []
    
    # إذا تم تحديد قسم معين
    if department_id:
        filter_conditions = [
            Employee.status == 'active',
            Employee.department_id == department_id
        ]
    else:
        filter_conditions = [Employee.status == 'active']
    
    # استعلام الموظفين الغائبين
    absent_records = db.session.query(
        Attendance, Employee, Department
    ).join(
        Employee, Employee.id == Attendance.employee_id
    ).join(
        Department, Department.id == Employee.department_id
    ).filter(
        Attendance.date == selected_date,
        Attendance.status.in_(['absent', 'leave', 'sick']),
        *filter_conditions
    ).all()
    
    # تجهيز بيانات الموظفين الغائبين
    for attendance, employee, department in absent_records:
        absent_employees.append({
            'id': employee.id,
            'name': employee.name,
            'employee_id': employee.employee_id,
            'department': department.name,
            'department_id': department.id,
            'status': attendance.status,
            'notes': attendance.notes
        })
    
    # تجهيز المعلومات حسب القسم
    absent_by_department = {}
    for emp in absent_employees:
        dept_id = emp['department_id']
        if dept_id not in absent_by_department:
            absent_by_department[dept_id] = {
                'name': emp['department'],
                'employees': []
            }
        absent_by_department[dept_id]['employees'].append(emp)
    
    # تقديم الصفحة
    return render_template(
        'attendance/enhanced_dashboard.html',
        selected_date=selected_date,
        departments=departments,
        selected_department_id=int(department_id) if department_id else None,
        total_employees=total_employees,
        total_present=total_present,
        total_absent=total_absent,
        total_leave=total_leave,
        total_sick=total_sick,
        total_missing=total_missing,
        department_stats=department_stats,
        absent_employees=absent_employees,
        absent_by_department=absent_by_department
    )

@attendance_dashboard_bp.route('/data')
@login_required
@module_access_required(Module.ATTENDANCE)
def dashboard_data():
    """API لبيانات لوحة المعلومات"""
    date_str = request.args.get('date')
    
    try:
        if date_str:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            selected_date = datetime.now().date()
    except ValueError:
        selected_date = datetime.now().date()
    
    # الحصول على بيانات الإحصائيات
    departments = Department.query.all()
    
    # بيانات للرسم البياني
    department_names = []
    present_counts = []
    absent_counts = []
    leave_counts = []
    sick_counts = []
    missing_counts = []
    
    for dept in departments:
        # إضافة اسم القسم
        department_names.append(dept.name)
        
        # عدد الموظفين في القسم
        employees_count = Employee.query.filter_by(
            department_id=dept.id, 
            status='active'
        ).count()
        
        # عدد الحاضرين
        present_count = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.department_id == dept.id,
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'present'
        ).scalar() or 0
        
        # عدد الغائبين
        absent_count = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.department_id == dept.id,
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'absent'
        ).scalar() or 0
        
        # عدد الإجازات
        leave_count = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.department_id == dept.id,
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'leave'
        ).scalar() or 0
        
        # عدد المرضي
        sick_count = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.department_id == dept.id,
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'sick'
        ).scalar() or 0
        
        # إجمالي السجلات
        total_records = present_count + absent_count + leave_count + sick_count
        
        # حساب عدد الموظفين بدون تسجيل
        missing_count = employees_count - total_records
        if missing_count < 0:
            missing_count = 0
        
        # إضافة الإحصائيات للقوائم
        present_counts.append(present_count)
        absent_counts.append(absent_count)
        leave_counts.append(leave_count)
        sick_counts.append(sick_count)
        missing_counts.append(missing_count)
    
    # تجهيز البيانات للرسم البياني
    chart_data = {
        'labels': department_names,
        'datasets': [
            {
                'label': 'حاضر',
                'data': present_counts,
                'backgroundColor': 'rgba(40, 167, 69, 0.7)',
                'borderColor': 'rgb(40, 167, 69)',
                'borderWidth': 1
            },
            {
                'label': 'غائب',
                'data': absent_counts,
                'backgroundColor': 'rgba(220, 53, 69, 0.7)',
                'borderColor': 'rgb(220, 53, 69)',
                'borderWidth': 1
            },
            {
                'label': 'إجازة',
                'data': leave_counts,
                'backgroundColor': 'rgba(255, 193, 7, 0.7)',
                'borderColor': 'rgb(255, 193, 7)',
                'borderWidth': 1
            },
            {
                'label': 'مرضي',
                'data': sick_counts,
                'backgroundColor': 'rgba(23, 162, 184, 0.7)',
                'borderColor': 'rgb(23, 162, 184)',
                'borderWidth': 1
            },
            {
                'label': 'غير مسجل',
                'data': missing_counts,
                'backgroundColor': 'rgba(108, 117, 125, 0.7)',
                'borderColor': 'rgb(108, 117, 125)',
                'borderWidth': 1
            }
        ]
    }
    
    return jsonify(chart_data)

@attendance_dashboard_bp.route('/export-excel')
@login_required
@module_access_required(Module.ATTENDANCE)
def export_excel():
    """تصدير بيانات الحضور لملف إكسل مع داشبورد ورسم بياني"""
    date_str = request.args.get('date')
    department_id = request.args.get('department_id')
    
    try:
        if date_str:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            selected_date = datetime.now().date()
    except ValueError:
        selected_date = datetime.now().date()
    
    # استخراج بيانات الحضور التفصيلية
    detailed_query = db.session.query(
        Employee.name.label('اسم الموظف'),
        Employee.employee_id.label('الرقم الوظيفي'),
        Department.name.label('القسم'),
        Attendance.status.label('الحالة'),
        Attendance.check_in.label('وقت الحضور'),
        Attendance.check_out.label('وقت الانصراف'),
        Attendance.notes.label('ملاحظات')
    ).join(
        Employee, Employee.id == Attendance.employee_id
    ).join(
        Department, Department.id == Employee.department_id
    ).filter(
        Attendance.date == selected_date,
        Employee.status == 'active'
    )
    
    # إضافة فلتر القسم إذا تم تحديده
    if department_id:
        detailed_query = detailed_query.filter(Department.id == department_id)
    
    # تنفيذ الاستعلام التفصيلي
    detailed_results = detailed_query.all()
    
    # إنشاء DataFrame من النتائج التفصيلية
    detailed_df = pd.DataFrame(detailed_results)
    
    # معالجة الحالة
    def format_status(status):
        if status == 'present':
            return 'حاضر'
        elif status == 'absent':
            return 'غائب'
        elif status == 'leave':
            return 'إجازة'
        elif status == 'sick':
            return 'مرضي'
        else:
            return status
    
    if 'الحالة' in detailed_df.columns and not detailed_df.empty:
        detailed_df['الحالة'] = detailed_df['الحالة'].apply(format_status)
    
    # معالجة أوقات الحضور والانصراف
    for col in ['وقت الحضور', 'وقت الانصراف']:
        if col in detailed_df.columns and not detailed_df.empty:
            detailed_df[col] = detailed_df[col].apply(lambda x: x.strftime('%H:%M') if x else '-')
    
    # استخراج إحصائيات الحضور حسب القسم
    departments = Department.query.all()
    
    # إنشاء بيانات الداشبورد
    dashboard_data = []
    
    # حساب الإحصائيات لكل قسم
    for dept in departments:
        # عدد الموظفين في القسم
        employees_count = Employee.query.filter_by(
            department_id=dept.id, 
            status='active'
        ).count()
        
        # عدد الحاضرين
        present_count = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.department_id == dept.id,
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'present'
        ).scalar() or 0
        
        # عدد الغائبين
        absent_count = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.department_id == dept.id,
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'absent'
        ).scalar() or 0
        
        # عدد الإجازات
        leave_count = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.department_id == dept.id,
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'leave'
        ).scalar() or 0
        
        # عدد المرضي
        sick_count = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.department_id == dept.id,
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'sick'
        ).scalar() or 0
        
        # غير مسجلين
        total_records = present_count + absent_count + leave_count + sick_count
        missing_count = employees_count - total_records
        if missing_count < 0:
            missing_count = 0
        
        # نسبة الحضور
        if employees_count > 0:
            present_percentage = round((present_count / employees_count) * 100, 1)
            absent_percentage = round((absent_count / employees_count) * 100, 1)
        else:
            present_percentage = 0
            absent_percentage = 0
        
        # إضافة بيانات القسم
        dashboard_data.append({
            'القسم': dept.name,
            'عدد الموظفين': employees_count,
            'الحاضرون': present_count,
            'الغائبون': absent_count,
            'إجازة': leave_count,
            'مرضي': sick_count,
            'غير مسجلين': missing_count,
            'نسبة الحضور %': present_percentage,
            'نسبة الغياب %': absent_percentage
        })
    
    # إنشاء DataFrame للداشبورد
    dashboard_df = pd.DataFrame(dashboard_data)
    
    # تحديد اسم الملف
    date_str = selected_date.strftime('%Y-%m-%d')
    filename = f"تقرير_الحضور_{date_str}.xlsx"
    
    # إنشاء ملف مؤقت
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
        # حفظ البيانات إلى الملف
        with pd.ExcelWriter(temp_file.name, engine='openpyxl') as writer:
            # إضافة ورقة داشبورد
            dashboard_df.to_excel(writer, sheet_name='لوحة الإحصائيات', index=False)
            
            # إضافة ورقة البيانات التفصيلية
            if not detailed_df.empty:
                detailed_df.to_excel(writer, sheet_name='بيانات الحضور', index=False)
            
            # الوصول إلى الكائن workbook
            workbook = writer.book
            
            # إضافة ورقة الرسم البياني
            chart_sheet = workbook.create_sheet(title='الرسم البياني')
            
            # إضافة بيانات للرسم البياني
            for r, row in enumerate(dashboard_data, start=1):
                chart_sheet.cell(row=r, column=1, value=row['القسم'])
                chart_sheet.cell(row=r, column=2, value=row['الحاضرون'])
                chart_sheet.cell(row=r, column=3, value=row['الغائبون'])
                chart_sheet.cell(row=r, column=4, value=row['إجازة'])
                chart_sheet.cell(row=r, column=5, value=row['مرضي'])
            
            # إضافة العناوين
            chart_sheet.cell(row=1, column=1, value='القسم')
            chart_sheet.cell(row=1, column=2, value='الحاضرون')
            chart_sheet.cell(row=1, column=3, value='الغائبون')
            chart_sheet.cell(row=1, column=4, value='إجازة')
            chart_sheet.cell(row=1, column=5, value='مرضي')
            
            # تنسيق ورقة الإحصائيات
            dashboard_sheet = writer.sheets['لوحة الإحصائيات']
            for col in dashboard_sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                dashboard_sheet.column_dimensions[column].width = adjusted_width
            
            # إضافة إطار وتلوين للعناوين
            for col_num, column_title in enumerate(dashboard_df.columns, 1):
                cell = dashboard_sheet.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
                
            # تلوين الخلايا حسب القيمة (للنسب المئوية)
            for row_num, row in enumerate(dashboard_df.iterrows(), 2):
                # الحصول على مؤشر العمود
                present_col = dashboard_df.columns.get_indexer(['نسبة الحضور %'])[0] + 1
                absent_col = dashboard_df.columns.get_indexer(['نسبة الغياب %'])[0] + 1
                
                present_cell = dashboard_sheet.cell(row=row_num, column=present_col)
                absent_cell = dashboard_sheet.cell(row=row_num, column=absent_col)
                
                present_value = row[1]['نسبة الحضور %']
                
                # تلوين نسبة الحضور
                if present_value >= 90:
                    present_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif present_value >= 70:
                    present_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                else:
                    present_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            # إنشاء مخطط شريطي (Bar Chart)
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "توزيع الحضور حسب الأقسام"
            chart.y_axis.title = "عدد الموظفين"
            chart.x_axis.title = "القسم"
            
            # تحديد نطاق البيانات
            rows_count = len(dashboard_data) + 1  # +1 للعناوين
            
            # إضافة البيانات للمخطط
            data = Reference(chart_sheet, min_col=2, min_row=1, max_row=rows_count, max_col=5)
            cats = Reference(chart_sheet, min_col=1, min_row=2, max_row=rows_count)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # إضافة المخطط إلى الورقة
            chart_sheet.add_chart(chart, "A10")
            
        # إرسال الملف كملف للتحميل
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )