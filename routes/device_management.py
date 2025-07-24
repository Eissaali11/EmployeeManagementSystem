from flask import Blueprint, render_template, request, flash, redirect, url_for, jsonify, Response
from flask_login import current_user
from sqlalchemy import func, and_, or_
from models import db, MobileDevice, Employee, Department, DeviceAssignment, ImportedPhoneNumber
from datetime import datetime
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

device_management_bp = Blueprint('device_management', __name__)

@device_management_bp.route('/')
def index():
    """عرض صفحة قائمة الأجهزة المحمولة مع الفلاتر والإحصائيات"""
    try:
        # استلام الفلاتر من الطلب
        department_filter = request.args.get('department', '')
        brand_filter = request.args.get('brand', '')
        status_filter = request.args.get('status', '')
        search_term = request.args.get('search', '')
        page = request.args.get('page', 1, type=int)
        
        # بناء الاستعلام
        query = MobileDevice.query
        
        # تطبيق الفلاتر
        if department_filter:
            query = query.join(Employee).join(Employee.departments).filter(Department.id == department_filter)
        
        if brand_filter:
            query = query.filter(MobileDevice.device_brand.like(f'%{brand_filter}%'))
        
        # تطبيق فلتر الحالة بناءً على جدول DeviceAssignment
        device_ids_assigned = []
        device_ids_available = []
        
        if status_filter:
            if status_filter == 'assigned':
                # الحصول على معرفات الأجهزة المربوطة
                assigned_assignments = DeviceAssignment.query.filter_by(is_active=True).filter(DeviceAssignment.device_id.isnot(None)).all()
                device_ids_assigned = [assignment.device_id for assignment in assigned_assignments]
                if device_ids_assigned:
                    query = query.filter(MobileDevice.id.in_(device_ids_assigned))
                else:
                    # لا توجد أجهزة مربوطة
                    query = query.filter(MobileDevice.id == -1)  # فلتر فارغ
                    
            elif status_filter == 'available':
                # الحصول على معرفات الأجهزة المربوطة لاستبعادها
                assigned_assignments = DeviceAssignment.query.filter_by(is_active=True).filter(DeviceAssignment.device_id.isnot(None)).all()
                device_ids_assigned = [assignment.device_id for assignment in assigned_assignments]
                if device_ids_assigned:
                    query = query.filter(~MobileDevice.id.in_(device_ids_assigned))
                # إذا لم توجد أجهزة مربوطة، فجميع الأجهزة متاحة
                    
            elif status_filter == 'no_phone':
                query = query.filter(or_(MobileDevice.phone_number.is_(None), MobileDevice.phone_number == ''))
            elif status_filter == 'with_phone':
                query = query.filter(and_(MobileDevice.phone_number.isnot(None), MobileDevice.phone_number != ''))
        
        if search_term:
            query = query.filter(
                or_(
                    MobileDevice.imei.like(f'%{search_term}%'),
                    MobileDevice.phone_number.like(f'%{search_term}%')
                )
            )
        
        # ترتيب وصفحة
        devices = query.order_by(MobileDevice.created_at.desc()).all()
        
        # إضافة معلومات الربط لكل جهاز
        for device in devices:
            # البحث عن الربط النشط للجهاز
            active_assignment = DeviceAssignment.query.filter_by(
                device_id=device.id, 
                is_active=True
            ).first()
            
            if active_assignment:
                device.current_assignment = active_assignment
                device.assigned_employee = active_assignment.employee
                device.assigned_sim = active_assignment.sim_card
                device.is_assigned = True
            else:
                device.current_assignment = None
                device.assigned_employee = None
                device.assigned_sim = None
                device.is_assigned = False
        
        # حساب الإحصائيات من جدول DeviceAssignment
        total_devices = MobileDevice.query.count()
        assigned_assignments = DeviceAssignment.query.filter_by(is_active=True).filter(DeviceAssignment.device_id.isnot(None)).count()
        available_devices = total_devices - assigned_assignments
        assigned_devices = assigned_assignments
        
        # إحصائيات الأرقام
        devices_with_phone = MobileDevice.query.filter(and_(MobileDevice.phone_number.isnot(None), MobileDevice.phone_number != '')).count()
        devices_without_phone = MobileDevice.query.filter(or_(MobileDevice.phone_number.is_(None), MobileDevice.phone_number == '')).count()
        
        # إحصائيات العلامات التجارية
        iphone_count = MobileDevice.query.filter(MobileDevice.device_brand.like('%iPhone%')).count()
        samsung_count = MobileDevice.query.filter(MobileDevice.device_brand.like('%Samsung%')).count()
        
        stats = {
            'total_devices': total_devices,
            'available_devices': available_devices,
            'assigned_devices': assigned_devices,
            'devices_with_phone': devices_with_phone,
            'devices_without_phone': devices_without_phone,
            'iphone_count': iphone_count,
            'samsung_count': samsung_count
        }
        
        # جلب قائمة الأقسام والموظفين
        departments = Department.query.all()
        employees = Employee.query.filter_by(status='نشط').all()
        
        return render_template('device_management/index.html',
                             devices=devices,
                             stats=stats,
                             departments=departments,
                             employees=employees,
                             department_filter=department_filter,
                             brand_filter=brand_filter,
                             status_filter=status_filter,
                             search_term=search_term)
    
    except Exception as e:
        flash(f'حدث خطأ في تحميل البيانات: {str(e)}', 'error')
        return render_template('device_management/index.html', devices=[], stats={}, departments=[])

@device_management_bp.route('/create', methods=['GET', 'POST'])
def create():
    """إضافة جهاز محمول جديد"""
    if request.method == 'POST':
        try:
            # استلام البيانات من النموذج
            imei = request.form.get('imei')
            brand = request.form.get('brand')
            model = request.form.get('model')
            phone_number = request.form.get('phone_number')
            serial_number = request.form.get('serial_number')
            purchase_date = request.form.get('purchase_date')
            warranty_expiry = request.form.get('warranty_expiry')
            description = request.form.get('description')
            
            # التحقق من صحة IMEI
            if not imei or len(imei) != 15:
                flash('رقم IMEI يجب أن يكون 15 رقم بالضبط', 'error')
                return render_template('device_management/create.html')
            
            # التحقق من عدم تكرار IMEI
            existing_device = MobileDevice.query.filter_by(imei=imei).first()
            if existing_device:
                flash('رقم IMEI موجود مسبقاً في النظام', 'error')
                return render_template('device_management/create.html')
            
            # إنشاء الجهاز الجديد
            device = MobileDevice()
            device.imei = imei
            device.device_brand = brand
            device.device_model = model
            device.phone_number = phone_number or None
            device.notes = description
            
            # إضافة التواريخ إذا تم إدخالها
            if purchase_date:
                device.assigned_date = datetime.strptime(purchase_date, '%Y-%m-%d')
            
            db.session.add(device)
            db.session.commit()
            
            # تسجيل العملية في السجل
            # audit_log('device_created', f'تم إضافة جهاز جديد: {brand} {model} - IMEI: {imei}')
            
            flash('تم إضافة الجهاز بنجاح', 'success')
            return redirect(url_for('device_management.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ في إضافة الجهاز: {str(e)}', 'error')
    
    return render_template('device_management/create.html')

@device_management_bp.route('/edit/<int:device_id>', methods=['GET', 'POST'])
def edit(device_id):
    """تعديل بيانات جهاز محمول"""
    device = MobileDevice.query.get_or_404(device_id)
    
    if request.method == 'POST':
        try:
            # التحقق من IMEI الجديد
            new_imei = request.form.get('imei')
            if not new_imei or len(new_imei) != 15:
                flash('رقم IMEI يجب أن يكون 15 رقم بالضبط', 'error')
                return render_template('device_management/edit.html', device=device)
            
            # التحقق من عدم تكرار IMEI
            if new_imei != device.imei:
                existing_device = MobileDevice.query.filter_by(imei=new_imei).first()
                if existing_device:
                    flash('رقم IMEI موجود مسبقاً في النظام', 'error')
                    return render_template('device_management/edit.html', device=device)
            
            # تحديث البيانات
            device.imei = new_imei
            device.device_brand = request.form.get('brand')
            device.device_model = request.form.get('model')
            device.phone_number = request.form.get('phone_number') or None
            device.notes = request.form.get('description')
            device.updated_at = datetime.utcnow()
            
            db.session.commit()
            
            # تسجيل العملية
            # audit_log('device_updated', f'تم تحديث الجهاز: {device.device_brand} {device.device_model} - IMEI: {device.imei}')
            
            flash('تم تحديث بيانات الجهاز بنجاح', 'success')
            return redirect(url_for('device_management.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ في تحديث الجهاز: {str(e)}', 'error')
    
    return render_template('device_management/edit.html', device=device)

@device_management_bp.route('/delete/<int:device_id>', methods=['POST'])
def delete(device_id):
    """حذف جهاز محمول"""
    try:
        device = MobileDevice.query.get_or_404(device_id)
        
        # التحقق من أن الجهاز غير مربوط بموظف
        if device.employee_id:
            flash('لا يمكن حذف جهاز مربوط بموظف. يرجى فك الربط أولاً', 'error')
            return redirect(url_for('device_management.index'))
        
        device_info = f'{device.device_brand} {device.device_model} - IMEI: {device.imei}'
        
        db.session.delete(device)
        db.session.commit()
        
        # تسجيل العملية
        # audit_log('device_deleted', f'تم حذف الجهاز: {device_info}')
        
        flash('تم حذف الجهاز بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ في حذف الجهاز: {str(e)}', 'error')
    
    return redirect(url_for('device_management.index'))

@device_management_bp.route('/assign/<int:device_id>', methods=['POST'])
def assign_to_employee(device_id):
    """ربط الجهاز بموظف"""
    try:
        device = MobileDevice.query.get_or_404(device_id)
        employee_id = request.form.get('employee_id')
        
        if not employee_id:
            flash('يرجى اختيار موظف', 'error')
            return redirect(url_for('device_management.index'))
        
        employee = Employee.query.get_or_404(employee_id)
        
        # التحقق من أن الجهاز متاح
        if device.employee_id:
            flash('الجهاز مربوط بموظف آخر بالفعل', 'error')
            return redirect(url_for('device_management.index'))
        
        # إنشاء سجل جديد في DeviceAssignment
        assignment = DeviceAssignment()
        assignment.device_id = device_id
        assignment.employee_id = employee_id
        assignment.assigned_date = datetime.utcnow()
        assignment.assigned_by = current_user.id if current_user.is_authenticated else None
        assignment.is_active = True
        assignment.assignment_type = 'مباشر'  # ربط مباشر من صفحة إدارة الأجهزة
        
        # ربط الجهاز
        device.employee_id = employee_id
        device.status = 'مرتبط'
        device.assigned_date = datetime.utcnow()
        device.updated_at = datetime.utcnow()
        
        db.session.add(assignment)
        db.session.commit()
        
        # تسجيل العملية
        from utils.audit_logger import log_activity
        log_activity(
            action='device_assigned',
            entity_type='DeviceAssignment',
            entity_id=assignment.id,
            details=f'تم ربط الجهاز {device.device_brand} {device.device_model} بالموظف {employee.name}'
        )
        
        flash(f'تم ربط الجهاز بالموظف {employee.name} بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ في ربط الجهاز: {str(e)}', 'error')
    
    return redirect(url_for('device_management.index'))

@device_management_bp.route('/unassign/<int:device_id>', methods=['POST'])
def unassign_from_employee(device_id):
    """فك ربط الجهاز من الموظف"""
    try:
        print(f"DEBUG: بدء عملية فك ربط الجهاز {device_id}")
        device = MobileDevice.query.get_or_404(device_id)
        
        # العثور على جميع الربطات النشطة لهذا الجهاز
        active_assignments = DeviceAssignment.query.filter(
            DeviceAssignment.device_id == device_id,
            DeviceAssignment.is_active == True
        ).all()
        
        print(f"DEBUG: عدد الربطات النشطة الموجودة: {len(active_assignments)}")
        
        if not active_assignments:
            flash('الجهاز غير مربوط بأي موظف', 'warning')
            return redirect(url_for('device_management.index'))
        
        employee_name = 'غير معروف'
        
        # إلغاء جميع الربطات النشطة
        for assignment in active_assignments:
            if assignment.employee:
                employee_name = assignment.employee.name
            
            print(f"DEBUG: إلغاء الربط {assignment.id} للموظف {employee_name}")
            assignment.is_active = False
            assignment.unassigned_date = datetime.utcnow()
            assignment.unassigned_by = current_user.id if current_user.is_authenticated else None
            
            # إذا كان هناك رقم مربوط، قم بفك ربطه أيضاً
            if assignment.sim_card_id:
                sim_card = ImportedPhoneNumber.query.get(assignment.sim_card_id)
                if sim_card:
                    sim_card.employee_id = None
                    sim_card.is_used = False
                    print(f"DEBUG: تم فك ربط الرقم {sim_card.phone_number}")
        
        # فك ربط الجهاز في الجدول الأساسي
        device.employee_id = None
        device.status = 'متاح'
        device.updated_at = datetime.utcnow()
        print(f"DEBUG: تم تحديث حالة الجهاز إلى متاح")
        
        db.session.commit()
        print(f"DEBUG: تم حفظ التغييرات في قاعدة البيانات")
        
        # تسجيل العملية
        from utils.audit_logger import log_activity
        log_activity(
            action='device_unassigned',
            entity_type='DeviceAssignment',
            entity_id=device_id,
            details=f'تم فك ربط الجهاز {device.device_brand} {device.device_model} من الموظف {employee_name}'
        )
        
        flash(f'تم فك ربط الجهاز من الموظف {employee_name} بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        print(f"DEBUG: خطأ في فك الربط: {str(e)}")
        flash(f'حدث خطأ في فك ربط الجهاز: {str(e)}', 'error')
    
    return redirect(url_for('device_management.index'))

@device_management_bp.route('/export-excel')
def export_excel():
    """تصدير بيانات الأجهزة إلى Excel بناءً على الفلاتر المطبقة"""
    try:
        # جلب الفلاتر من parameters
        department_filter = request.args.get('department', '')
        brand_filter = request.args.get('brand', '')
        status_filter = request.args.get('status', '')
        search_term = request.args.get('search', '')
        
        # بناء الاستعلام مع نفس منطق الفلترة من دالة index
        query = MobileDevice.query
        
        # فلتر القسم
        if department_filter:
            # جلب موظفي القسم المحدد
            employees_in_dept = Employee.query.join(Employee.departments).filter(Department.id == department_filter).all()
            employee_ids = [emp.id for emp in employees_in_dept]
            
            # فلترة الأجهزة المربوطة بموظفي هذا القسم
            if employee_ids:
                assigned_device_ids = [assignment.device_id for assignment in 
                                     DeviceAssignment.query.filter_by(is_active=True)
                                     .filter(DeviceAssignment.employee_id.in_(employee_ids))
                                     .filter(DeviceAssignment.device_id.isnot(None)).all()]
                if assigned_device_ids:
                    query = query.filter(MobileDevice.id.in_(assigned_device_ids))
                else:
                    query = query.filter(MobileDevice.id == -1)  # فلتر فارغ
            else:
                query = query.filter(MobileDevice.id == -1)  # فلتر فارغ
        
        # فلتر العلامة التجارية
        if brand_filter:
            query = query.filter(MobileDevice.device_brand.like(f'%{brand_filter}%'))
        
        # فلتر الحالة
        if status_filter:
            if status_filter == 'assigned':
                # الأجهزة المربوطة
                assigned_assignments = DeviceAssignment.query.filter_by(is_active=True).filter(DeviceAssignment.device_id.isnot(None)).all()
                device_ids_assigned = [assignment.device_id for assignment in assigned_assignments]
                if device_ids_assigned:
                    query = query.filter(MobileDevice.id.in_(device_ids_assigned))
                else:
                    query = query.filter(MobileDevice.id == -1)
                    
            elif status_filter == 'available':
                # الأجهزة المتاحة
                assigned_assignments = DeviceAssignment.query.filter_by(is_active=True).filter(DeviceAssignment.device_id.isnot(None)).all()
                device_ids_assigned = [assignment.device_id for assignment in assigned_assignments]
                if device_ids_assigned:
                    query = query.filter(~MobileDevice.id.in_(device_ids_assigned))
                    
            elif status_filter == 'no_phone':
                query = query.filter(or_(MobileDevice.phone_number.is_(None), MobileDevice.phone_number == ''))
            elif status_filter == 'with_phone':
                query = query.filter(and_(MobileDevice.phone_number.isnot(None), MobileDevice.phone_number != ''))
        
        # فلتر البحث
        if search_term:
            search_filter = or_(
                MobileDevice.imei.like(f'%{search_term}%'),
                MobileDevice.device_brand.like(f'%{search_term}%'),
                MobileDevice.device_model.like(f'%{search_term}%'),
                MobileDevice.phone_number.like(f'%{search_term}%')
            )
            query = query.filter(search_filter)
        
        # تنفيذ الاستعلام
        devices = query.order_by(MobileDevice.created_at.desc()).all()
        
        # إنشاء workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "الأجهزة المحمولة"
        
        # إعداد الرؤوس
        headers = ['IMEI', 'العلامة التجارية', 'الموديل', 'رقم الهاتف', 'الحالة', 'الموظف المرتبط', 'القسم', 'نوع الربط', 'تاريخ الربط', 'ملاحظات']
        ws.append(headers)
        
        # تنسيق الرؤوس
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # إضافة البيانات
        for device in devices:
            # إضافة معلومات الربط النشط
            current_assignment = DeviceAssignment.query.filter_by(
                device_id=device.id, is_active=True
            ).first()
            
            employee_name = ''
            department_name = ''
            assignment_type = ''
            assigned_date = ''
            status = 'متاح'
            
            if current_assignment:
                status = 'مربوط'
                assigned_date = current_assignment.assignment_date.strftime('%Y-%m-%d') if current_assignment.assignment_date else ''
                
                if current_assignment.assignment_target_type == 'employee' and current_assignment.employee:
                    employee_name = current_assignment.employee.name
                    # جلب أقسام الموظف
                    departments = [dept.name for dept in current_assignment.employee.departments]
                    department_name = ', '.join(departments) if departments else ''
                elif current_assignment.assignment_target_type == 'department' and current_assignment.department:
                    department_name = current_assignment.department.name
                    employee_name = f'قسم: {department_name}'
                
                # نوع الربط
                if current_assignment.assignment_type == 'device_only':
                    assignment_type = 'جهاز فقط'
                elif current_assignment.assignment_type == 'sim_only':
                    assignment_type = 'رقم فقط'
                elif current_assignment.assignment_type == 'device_and_sim':
                    assignment_type = 'جهاز ورقم'
            
            row_data = [
                device.imei,
                device.device_brand or '',
                device.device_model or '',
                device.phone_number or '',
                status,
                employee_name,
                department_name,
                assignment_type,
                assigned_date,
                device.notes or ''
            ]
            ws.append(row_data)
        
        # ضبط عرض الأعمدة
        column_widths = [20, 15, 20, 15, 12, 20, 15, 12, 15, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
        
        # حفظ في memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # إنشاء اسم الملف بناءً على الفلاتر (بالإنجليزية لتجنب مشاكل الترميز)
        filename = "mobile_devices"
        if department_filter:
            dept = Department.query.get(department_filter)
            if dept:
                # تحويل الاسم العربي إلى شكل آمن
                safe_name = dept.name.replace(' ', '_').replace('/', '_')
                filename += f"_dept_{dept.id}"
        if brand_filter:
            safe_brand = brand_filter.replace(' ', '_').replace('/', '_')
            filename += f"_{safe_brand}"
        if status_filter:
            status_names = {
                'assigned': 'assigned',
                'available': 'available', 
                'no_phone': 'no_phone',
                'with_phone': 'with_phone'
            }
            filename += f"_{status_names.get(status_filter, status_filter)}"
        
        filename += ".xlsx"
        
        # ترميز اسم الملف بشكل آمن للمتصفحات
        from urllib.parse import quote
        safe_filename = quote(filename.encode('utf-8'))
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename*=UTF-8\'\'{safe_filename}'
            }
        )
        
    except Exception as e:
        flash(f'حدث خطأ في تصدير البيانات: {str(e)}', 'error')
        return redirect(url_for('device_management.index'))

@device_management_bp.route('/devices-only')
def devices_only():
    """صفحة الأجهزة بدون أرقام هاتف فقط"""
    try:
        # جلب جميع الأجهزة بدون أرقام هاتف
        devices = MobileDevice.query.filter(
            or_(MobileDevice.phone_number.is_(None), MobileDevice.phone_number == '')
        ).order_by(MobileDevice.created_at.desc()).all()
        
        # جلب الأقسام للفلاتر
        departments = Department.query.order_by(Department.name).all()
        
        # إحصائيات خاصة بالأجهزة بدون أرقام
        total_devices_only = len(devices)
        available_devices_only = len([d for d in devices if d.employee_id is None])
        assigned_devices_only = len([d for d in devices if d.employee_id is not None])
        
        # إحصائيات العلامات التجارية للأجهزة بدون أرقام
        iphone_count = len([d for d in devices if d.device_brand and 'iPhone' in d.device_brand])
        samsung_count = len([d for d in devices if d.device_brand and 'Samsung' in d.device_brand])
        
        stats = {
            'total_devices': total_devices_only,
            'available_devices': available_devices_only,
            'assigned_devices': assigned_devices_only,
            'iphone_count': iphone_count,
            'samsung_count': samsung_count
        }
        
        return render_template('device_management/devices_only.html',
                             devices=devices,
                             departments=departments,
                             stats=stats)
                             
    except Exception as e:
        print(f"Error in devices_only: {str(e)}")
        flash('حدث خطأ في تحميل البيانات', 'error')
        return redirect(url_for('device_management.index'))

@device_management_bp.route('/import-excel', methods=['GET', 'POST'])
def import_excel():
    """استيراد الأجهزة من ملف Excel"""
    if request.method == 'POST':
        try:
            if 'excel_file' not in request.files:
                flash('لم يتم اختيار ملف', 'error')
                return render_template('device_management/import_excel.html')
            
            file = request.files['excel_file']
            if file.filename == '':
                flash('لم يتم اختيار ملف', 'error')
                return render_template('device_management/import_excel.html')
            
            # قراءة الملف
            df = pd.read_excel(file)
            
            imported_count = 0
            skipped_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # استخراج البيانات
                    imei = str(row.get('IMEI', '')).strip()
                    brand = str(row.get('العلامة التجارية', '')).strip()
                    model = str(row.get('الموديل', '')).strip()
                    phone_number = str(row.get('رقم الهاتف', '')).strip()
                    description = str(row.get('الوصف', '')).strip()
                    
                    # التحقق من البيانات الأساسية
                    if not imei or not brand or not model:
                        errors.append(f'الصف {index + 2}: IMEI أو العلامة التجارية أو الموديل مفقود')
                        continue
                    
                    if len(imei) != 15:
                        errors.append(f'الصف {index + 2}: رقم IMEI يجب أن يكون 15 رقم')
                        continue
                    
                    # التحقق من عدم وجود الجهاز
                    existing_device = MobileDevice.query.filter_by(imei=imei).first()
                    if existing_device:
                        skipped_count += 1
                        continue
                    
                    # إنشاء الجهاز الجديد
                    device = MobileDevice()
                    device.imei = imei
                    device.device_brand = brand
                    device.device_model = model
                    device.phone_number = phone_number if phone_number and phone_number != 'nan' else None
                    device.notes = description if description and description != 'nan' else None
                    
                    db.session.add(device)
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f'الصف {index + 2}: {str(e)}')
            
            if imported_count > 0:
                db.session.commit()
                # audit_log('devices_imported', f'تم استيراد {imported_count} جهاز من Excel')
            
            # إعداد رسائل النتائج
            messages = []
            if imported_count > 0:
                messages.append(f'تم استيراد {imported_count} جهاز بنجاح')
            if skipped_count > 0:
                messages.append(f'تم تخطي {skipped_count} جهاز موجود مسبقاً')
            if errors:
                messages.append(f'فشل في استيراد {len(errors)} جهاز')
                for error in errors[:5]:  # عرض أول 5 أخطاء فقط
                    messages.append(f'- {error}')
            
            if imported_count > 0:
                flash('\n'.join(messages), 'success')
            else:
                flash('\n'.join(messages), 'warning')
            
            return redirect(url_for('device_management.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ في معالجة الملف: {str(e)}', 'error')
    
    return render_template('device_management/import_excel.html')