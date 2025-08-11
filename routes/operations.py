from flask import Blueprint, render_template, request, flash, redirect, url_for, jsonify, current_app
from flask_login import login_required, current_user
from app import db
from models import (OperationRequest, OperationNotification, VehicleHandover, 
                   VehicleWorkshop, ExternalAuthorization, SafetyInspection, 
                   Vehicle, User, UserRole, Employee)
from datetime import datetime
from utils.audit_logger import log_audit



from datetime import datetime, timedelta
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app, send_file
from flask_login import login_required, current_user
from flask_wtf import FlaskForm
from werkzeug.utils import secure_filename
from sqlalchemy import extract, func, or_
from forms.vehicle_forms import VehicleAccidentForm, VehicleDocumentsForm
import os
import uuid
import io
import urllib.parse
import pandas as pd
from fpdf import FPDF
import base64
import uuid

from app import db
from models import (
        Vehicle, VehicleRental, VehicleWorkshop, VehicleWorkshopImage, 
        VehicleProject, VehicleHandover, VehicleHandoverImage, SystemAudit,
        VehiclePeriodicInspection, VehicleSafetyCheck, VehicleAccident, Employee,
        Department, ExternalAuthorization, Module, Permission, UserRole,
        VehicleExternalSafetyCheck, OperationRequest
)
from utils.audit_logger import log_activity
from utils.audit_logger import log_audit
from utils.whatsapp_message_generator import generate_whatsapp_url
from utils.vehicles_export import export_vehicle_pdf, export_workshop_records_pdf, export_vehicle_excel, export_workshop_records_excel
from utils.simple_pdf_generator import create_vehicle_handover_pdf as generate_complete_vehicle_report
from utils.vehicle_excel_report import generate_complete_vehicle_excel_report
# from utils.workshop_report import generate_workshop_report_pdf
# from utils.html_to_pdf import generate_pdf_from_template
# from utils.fpdf_arabic_report import generate_workshop_report_pdf_fpdf
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
import arabic_reshaper
from bidi.algorithm import get_display
# from utils.fpdf_handover_pdf import generate_handover_report_pdf
# ============ تأكد من وجود هذه الاستيرادات في أعلى الملف ============
from datetime import date
# =================================================================


operations_bp = Blueprint('operations', __name__, url_prefix='/operations')



def update_vehicle_state(vehicle_id):
    """
    الدالة المركزية الذكية لتحديد وتحديث الحالة النهائية للمركبة وسائقها.
    (نسخة معدلة لا تعتمد على حقل is_approved).
    تعتمد على حالة OperationRequest المرتبط لتحديد السجلات الرسمية.
    """
    try:
        vehicle = Vehicle.query.get(vehicle_id)
        if not vehicle:
            current_app.logger.warning(f"محاولة تحديث حالة لمركبة غير موجودة: ID={vehicle_id}")
            return

        # 1. فحص الحالات ذات الأولوية القصوى (تبقى كما هي)
        if vehicle.status == 'out_of_service':
            return

        active_accident = VehicleAccident.query.filter(VehicleAccident.vehicle_id == vehicle_id, VehicleAccident.accident_status != 'مغلق').first()
        in_workshop = VehicleWorkshop.query.filter(VehicleWorkshop.vehicle_id == vehicle_id, VehicleWorkshop.exit_date.is_(None)).first()

        # نحدد ما إذا كانت السيارة في حالة حرجة
        is_critical_state = bool(active_accident or in_workshop)

        if active_accident:
            vehicle.status = 'accident'
        elif in_workshop:
            vehicle.status = 'in_workshop'

        # 2. التحقق من الإيجار النشط
        active_rental = VehicleRental.query.filter_by(vehicle_id=vehicle_id, is_active=True).first()

        # ================== بداية المنطق الجديد لتحديد السجلات الرسمية ==================

        # 3. إنشاء استعلام فرعي لجلب ID لكل سجل handover له طلب موافقة.
        approved_handover_ids_subquery = db.session.query(
            OperationRequest.related_record_id
        ).filter(
            OperationRequest.operation_type == 'handover',
            OperationRequest.status == 'approved',
            OperationRequest.vehicle_id == vehicle_id
        ).subquery()

        # 4. إنشاء استعلام فرعي لجلب ID لكل سجل handover له طلب (بغض النظر عن حالته).
        all_handover_request_ids_subquery = db.session.query(
            OperationRequest.related_record_id
        ).filter(
            OperationRequest.operation_type == 'handover',
            OperationRequest.vehicle_id == vehicle_id
        ).subquery()

        # 5. بناء الاستعلام الأساسي الذي يختار السجلات "الرسمية" فقط.
        # السجل يعتبر رسمياً إذا تمت الموافقة عليه، أو إذا كان قديماً (ليس له طلب موافقة أصلاً).
        base_official_query = VehicleHandover.query.filter(
            VehicleHandover.vehicle_id == vehicle_id
        ).filter(
            or_(
                VehicleHandover.id.in_(approved_handover_ids_subquery),
                ~VehicleHandover.id.in_(all_handover_request_ids_subquery)
            )
        )

        # 6. الآن نستخدم هذا الاستعلام الرسمي للحصول على آخر عملية تسليم واستلام
        latest_delivery = base_official_query.filter(
            VehicleHandover.handover_type.in_(['delivery', 'تسليم'])
        ).order_by(VehicleHandover.created_at.desc()).first()

        latest_return = base_official_query.filter(
            VehicleHandover.handover_type.in_(['return', 'استلام', 'receive'])
        ).order_by(VehicleHandover.created_at.desc()).first()

        # =================== نهاية المنطق الجديد لتحديد السجلات الرسمية ===================

        # 7. تطبيق التحديثات على السائق والحالة بناءً على السجلات الرسمية فقط
        is_currently_handed_out = False
        if latest_delivery:
            if not latest_return or latest_delivery.created_at > latest_return.created_at:
                is_currently_handed_out = True

        if is_currently_handed_out:
            # السيناريو (أ): السيارة مسلّمة حالياً (بناءً على سجل معتمد)
            vehicle.driver_name = latest_delivery.person_name
            # تحديث الحالة فقط إذا لم تكن السيارة في حالة حرجة (ورشة/حادث)
            if not is_critical_state:
                vehicle.status = 'rented' if active_rental else 'in_project'
        else:
            # السيناريو (ب): السيارة متاحة (بناءً على سجل معتمد)
            vehicle.driver_name = None
            # تحديث الحالة فقط إذا لم تكن السيارة في حالة حرجة
            if not is_critical_state:
                vehicle.status = 'rented' if active_rental else 'available'

        db.session.commit()

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"خطأ في دالة update_vehicle_state لـ vehicle_id {vehicle_id}: {str(e)}")



# def update_vehicle_state(vehicle_id):
#     """
#     الدالة المركزية الذكية لتحديد وتحديث الحالة النهائية للمركبة وسائقها
#     بناءً على هرم أولويات الحالات (ورشة > إيجار > تسليم > متاحة).
#     """
#     try:
#         vehicle = Vehicle.query.get(vehicle_id)
#         if not vehicle:
#             return

#         # -- هرم أولوية الحالات (من الأعلى إلى الأدنى) --

#         # 1. حالة "خارج الخدمة": لها أعلى أولوية ولا تتغير تلقائياً
#         if vehicle.status == 'out_of_service':
#             # لا تفعل شيئاً، هذه الحالة لا تتغير إلا يدوياً
#             return

#         # 2. حالة "الحادث"
#         # يجب تعديل منطق الحادث بحيث تبقى الحالة accident حتى يتم إغلاق السجل
#         active_accident = VehicleAccident.query.filter(
#             VehicleAccident.vehicle_id == vehicle_id,
#             VehicleAccident.accident_status != 'مغلق' # نفترض أن 'مغلق' هي الحالة النهائية
#         ).first()
#         if active_accident:
#             vehicle.status = 'accident'
#             # (منطق السائق يبقى كما هو أدناه لأنه قد يكون هناك سائق وقت الحادث)

#         # 3. حالة "الورشة"
#         in_workshop = VehicleWorkshop.query.filter(
#             VehicleWorkshop.vehicle_id == vehicle_id,
#             VehicleWorkshop.exit_date.is_(None) # لا يزال في الورشة
#         ).first()
#         if in_workshop:
#             vehicle.status = 'in_workshop'
#             db.session.commit() # نحفظ الحالة وننهي لأنها ذات أولوية
#             return # إنهاء الدالة، لأن الورشة لها الأسبقية على الإيجار والتسليم

#         # --- إذا لم تكن السيارة في ورشة، ننتقل للحالات التشغيلية ---

#         # 4. حالة "مؤجرة"
#         active_rental = VehicleRental.query.filter(
#             VehicleRental.vehicle_id == vehicle_id,
#             VehicleRental.is_active == True
#         ).first()
#         if active_rental:
#             vehicle.status = 'rented'
#             # لا ننهي هنا، سنكمل لتحديد السائق

#         # 5. حالة "التسليم" و "متاحة" (نفس منطق الدالة السابقة)
#         latest_delivery = VehicleHandover.query.filter(
#             VehicleHandover.vehicle_id == vehicle_id,
#             VehicleHandover.handover_type.in_(['delivery', 'تسليم'])
#         ).order_by(VehicleHandover.handover_date.desc(), VehicleHandover.id.desc()).first()

#         latest_return = VehicleHandover.query.filter(
#             VehicleHandover.vehicle_id == vehicle_id,
#             VehicleHandover.handover_type.in_(['return', 'استلام', 'receive'])
#         ).order_by(VehicleHandover.handover_date.desc(), VehicleHandover.id.desc()).first()

#         is_currently_handed_out = False
#         if latest_delivery:
#             if not latest_return or latest_delivery.created_at > latest_return.created_at:
#                  is_currently_handed_out = True

#         if is_currently_handed_out:
#             # مسلمة لسائق
#             vehicle.driver_name = latest_delivery.person_name
#             # إذا لم تكن مؤجرة، فستكون في مشروع
#             if not active_rental: 
#                 vehicle.status = 'in_project'
#         else:
#             # تم استلامها أو لم تسلم أبداً
#             vehicle.driver_name = None
#             # إذا لم تكن مؤجرة، فستكون متاحة
#             if not active_rental:
#                 vehicle.status = 'available'

#         db.session.commit()

#     except Exception as e:
#         db.session.rollback()
#         current_app.logger.error(f"خطأ في تحديث حالة المركبة {vehicle_id}: {e}")



@operations_bp.route('/')
@login_required
def operations_dashboard():
    """لوحة إدارة العمليات الرئيسية"""
    
    # فحص صلاحية المدير
    if current_user.role != UserRole.ADMIN:
        flash('غير مسموح لك بالوصول لهذه الصفحة', 'danger')
        return redirect(url_for('dashboard.index'))
    
    # إحصائيات العمليات
    pending_operations = OperationRequest.query.filter_by(status='pending').count()
    under_review_operations = OperationRequest.query.filter_by(status='under_review').count()
    approved_operations = OperationRequest.query.filter_by(status='approved').count()
    rejected_operations = OperationRequest.query.filter_by(status='rejected').count()
    
    # العمليات المعلقة مرتبة بالأولوية والتاريخ
    pending_requests = OperationRequest.query.filter_by(status='pending').order_by(
        OperationRequest.priority.desc(),
        OperationRequest.requested_at.desc()
    ).limit(10).all()
    
    # الإشعارات غير المقروءة
    unread_notifications = OperationNotification.query.filter_by(
        user_id=current_user.id, 
        is_read=False
    ).count()
    
    stats = {
        'pending': pending_operations,
        'under_review': under_review_operations,
        'approved': approved_operations,
        'rejected': rejected_operations,
        'unread_notifications': unread_notifications
    }
    
    return render_template('operations/dashboard.html', 
                         stats=stats, 
                         pending_requests=pending_requests)

@operations_bp.route('/list')
@login_required
def operations_list():
    """قائمة جميع العمليات مع فلترة"""
    
    if current_user.role != UserRole.ADMIN:
        flash('غير مسموح لك بالوصول لهذه الصفحة', 'danger')
        return redirect(url_for('dashboard.index'))
    
    # فلترة العمليات
    status_filter = request.args.get('status', 'all')
    operation_type_filter = request.args.get('operation_type', 'all')
    priority_filter = request.args.get('priority', 'all')
    vehicle_search = request.args.get('vehicle_search', '').strip()
    
    query = OperationRequest.query
    
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    
    if operation_type_filter != 'all':
        query = query.filter_by(operation_type=operation_type_filter)
    
    if priority_filter != 'all':
        query = query.filter_by(priority=priority_filter)
    
    # البحث حسب السيارة
    if vehicle_search:
        # البحث في عنوان العملية أو الوصف حيث يتم ذكر رقم السيارة عادة
        query = query.filter(
            OperationRequest.title.contains(vehicle_search) |
            OperationRequest.description.contains(vehicle_search)
        )
    
    # ترتيب العمليات
    operations = query.order_by(
        OperationRequest.priority.desc(),
        OperationRequest.requested_at.desc()
    ).all()
    
    return render_template('operations/list.html', 
                         operations=operations,
                         status_filter=status_filter,
                         operation_type_filter=operation_type_filter,
                         priority_filter=priority_filter,
                         vehicle_search=vehicle_search)

@operations_bp.route('/<int:operation_id>')
@login_required
def view_operation(operation_id):
    """عرض تفاصيل العملية"""
    
    if current_user.role != UserRole.ADMIN:
        flash('غير مسموح لك بالوصول لهذه الصفحة', 'danger')
        return redirect(url_for('dashboard.index'))
    
    operation = OperationRequest.query.get_or_404(operation_id)
    related_record = operation.get_related_record()
    
    # توجيه عمليات الورشة إلى صفحة منفصلة
    if operation.operation_type == 'workshop_record':
        # جلب معلومات السائق الحالي من آخر تسليم
        driver_employee = None
        current_driver_info = None
        
        # البحث عن آخر تسليم للمركبة للحصول على السائق الحالي
        last_handover = VehicleHandover.query.filter_by(
            vehicle_id=operation.vehicle_id,
            handover_type='delivery'
        ).order_by(VehicleHandover.handover_date.desc()).first()
        
        if last_handover:
            current_driver_info = {
                'name': last_handover.person_name,
                'phone': last_handover.driver_phone_number,
                'residency_number': last_handover.driver_residency_number
            }
            
            # محاولة العثور على بيانات الموظف في النظام
            if last_handover.driver_residency_number:
                driver_employee = Employee.query.filter_by(
                    national_id=last_handover.driver_residency_number
                ).first()
            
            if not driver_employee and last_handover.person_name:
                driver_employee = Employee.query.filter_by(
                    name=last_handover.person_name
                ).first()
        
        # جلب الصور من قاعدة البيانات مباشرة إذا كانت العلاقة لا تعمل
        from sqlalchemy import text
        workshop_images = []
        if related_record and hasattr(related_record, 'id'):
            try:
                # استعلام مباشر لجلب الصور
                result = db.session.execute(
                    text("SELECT id, image_type, image_path, notes, uploaded_at FROM vehicle_workshop_images WHERE workshop_record_id = :workshop_id ORDER BY uploaded_at DESC"),
                    {'workshop_id': related_record.id}
                )
                workshop_images = [
                    {
                        'id': row[0],
                        'image_type': row[1], 
                        'image_path': row[2],
                        'notes': row[3],
                        'uploaded_at': row[4]
                    } 
                    for row in result
                ]
                current_app.logger.debug(f"تم جلب {len(workshop_images)} صور لسجل الورشة {related_record.id}")
            except Exception as e:
                current_app.logger.error(f"خطأ في جلب صور الورشة: {str(e)}")
        
        return render_template('operations/view_workshop.html', 
                             operation=operation,
                             related_record=related_record,
                             workshop_images=workshop_images,
                             driver_employee=driver_employee,
                             current_driver_info=current_driver_info)
    
    # جلب بيانات الموظف إذا كانت متاحة
    employee = None
    if related_record:
        # محاولة العثور على الموظف من خلال رقم الهوية أو الاسم
        if hasattr(related_record, 'driver_residency_number') and related_record.driver_residency_number:
            employee = Employee.query.filter_by(national_id=related_record.driver_residency_number).first()
        if not employee and hasattr(related_record, 'person_name') and related_record.person_name:
            employee = Employee.query.filter_by(name=related_record.person_name).first()
        if not employee and hasattr(related_record, 'driver_name') and related_record.driver_name:
            employee = Employee.query.filter_by(name=related_record.driver_name).first()
    
    # تحميل معلومات إضافية للطالب إذا كان موظفاً
    if operation.requester and hasattr(operation.requester, 'id'):
        try:
            # تحميل العلاقات بشكل صريح
            requester_employee = Employee.query.options(db.joinedload(Employee.departments)).get(operation.requester.id)
            if requester_employee:
                operation.requester = requester_employee
        except Exception as e:
            print(f"خطأ في تحميل بيانات الموظف الطالب: {e}")
    
    return render_template('operations/view.html', 
                         operation=operation,
                         related_record=related_record,
                         employee=employee)



@operations_bp.route('/<int:operation_id>/approve', methods=['POST'])
@login_required
def approve_operation(operation_id):
    """الموافقة على العملية مع تفعيل السجل المرتبط وتحديث حالة السيارة."""

    if current_user.role != UserRole.ADMIN:
        return jsonify({'success': False, 'message': 'غير مسموح لك بالقيام بهذا الإجراء'})

    operation = OperationRequest.query.get_or_404(operation_id)
    review_notes = request.form.get('review_notes', '').strip()

    try:
        # 1. تحديث حالة الطلب نفسه
        operation.status = 'approved'
        operation.reviewed_by = current_user.id
        operation.reviewed_at = datetime.utcnow()
        operation.review_notes = review_notes

        # 2. البحث عن السجل المرتبط وتفعيله إذا كان من نوع handover
        if operation.operation_type == 'handover':
            handover_record = VehicleHandover.query.get(operation.related_record_id)
            if handover_record:
                # === الخطوة الأهم: تحويل المسودة إلى سجل رسمي ===
                # handover_record.is_approved = True
                db.session.commit() # نحفظ التفعيل أولاً

                # === بعد التفعيل، نقوم بتحديث حالة السيارة الآن ===
                if operation.vehicle_id:
                    update_vehicle_state(operation.vehicle_id)
                    log_audit(
                        'update', 'vehicle_state', operation.vehicle_id,
                        f'تم تحديث حالة السيارة والسائق بعد الموافقة على الطلب #{operation.id}'
                    )
            else:
                current_app.logger.warning(f"لم يتم العثور على سجل Handover رقم {operation.related_record_id} للموافقة عليه.")

        # (يمكن إضافة منطق لأنواع أخرى من العمليات هنا مستقبلاً)

        # 3. إنشاء إشعار للطالب
        create_notification(
            operation_id=operation.id,
            user_id=operation.requested_by,
            notification_type='status_change',
            title=f'✅ تمت الموافقة على طلبك: {operation.title}',
            message=f'تمت الموافقة على طلبك من قبل {current_user.username}.'
        )

        db.session.commit()

        log_audit('approve', 'operation_request', operation.id, f'تمت الموافقة على العملية: {operation.title}')
        flash('تمت الموافقة على العملية وتحديث حالة المركبة بنجاح', 'success')
        return jsonify({'success': True, 'message': 'تمت الموافقة بنجاح'})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"خطأ في الموافقة على العملية #{operation_id}: {str(e)}")
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})



# @operations_bp.route('/<int:operation_id>/approve', methods=['POST'])
# @login_required
# def approve_operation(operation_id):
#     """الموافقة على العملية"""
    
#     if current_user.role != UserRole.ADMIN:
#         return jsonify({'success': False, 'message': 'غير مسموح'})
    
#     operation = OperationRequest.query.get_or_404(operation_id)
#     review_notes = request.form.get('review_notes', '').strip()
    
#     try:
#         # تحديث حالة العملية
#         operation.status = 'approved'
#         operation.reviewed_by = current_user.id
#         operation.reviewed_at = datetime.utcnow()
#         operation.review_notes = review_notes
        
#         # إنشاء إشعار للطالب
#         create_notification(
#             operation_id=operation.id,
#             user_id=operation.requested_by,
#             notification_type='status_change',
#             title=f'تمت الموافقة على العملية: {operation.title}',
#             message=f'تمت الموافقة على العملية من قبل {current_user.username}.\n{review_notes if review_notes else ""}'
#         )
        
#         db.session.commit()
        
#         # تحديث السائق الحالي وحالة السيارة إذا كانت العملية من نوع handover
#         if operation.operation_type == 'handover' and operation.vehicle_id:
#             try:
#                 from utils.vehicle_driver_utils import update_vehicle_driver_approved
#                 update_vehicle_driver_approved(operation.vehicle_id)
                
#                 # تحديث حالة السيارة بناءً على نوع التسليم/الاستلام
#                 handover_record = VehicleHandover.query.get(operation.related_record_id)
#                 if handover_record:
#                     vehicle = Vehicle.query.get(operation.vehicle_id)
#                     if vehicle and handover_record.handover_type:
#                         old_status = vehicle.status
                        
#                         # تحديد الحالة الجديدة بناءً على نوع العملية
#                         if handover_record.handover_type == 'return':  # استلام السيارة
#                             new_status = 'متاحة'
#                         elif handover_record.handover_type == 'delivery':  # تسليم السيارة
#                             new_status = 'في المشروع'
#                         else:
#                             new_status = None
                        
#                         if new_status and old_status != new_status:
#                             vehicle.status = new_status
#                             db.session.add(vehicle)
                            
#                             # تسجيل تغيير الحالة
#                             action_type = 'الاستلام' if handover_record.handover_type == 'return' else 'التسليم'
#                             log_audit(current_user.id, 'update', 'vehicle_status', vehicle.id,
#                                      f'تم تحديث حالة السيارة {vehicle.plate_number} إلى "{new_status}" بعد الموافقة على عملية {action_type}')
                            
#                             print(f"تم تحديث حالة السيارة {vehicle.plate_number} من '{old_status}' إلى '{new_status}' بعد الموافقة على عملية {action_type}")
                        
#             except Exception as e:
#                 print(f"خطأ في تحديث السائق وحالة السيارة بعد الموافقة: {e}")
#                 import traceback
#                 traceback.print_exc()
        
#         # تسجيل العملية
#         log_audit(current_user.id, 'approve', 'operation_request', operation.id, 
#                  f'تمت الموافقة على العملية: {operation.title}')
        
#         flash('تمت الموافقة على العملية بنجاح', 'success')
#         return jsonify({'success': True, 'message': 'تمت الموافقة بنجاح'})
        
#     except Exception as e:
#         db.session.rollback()
#         return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})


# @operations_bp.route('/<int:operation_id>/reject', methods=['POST'])
# @login_required
# def reject_operation(operation_id):
#     """رفض العملية مع حذف السجل المؤقت المرتبط بها."""

#     if current_user.role != UserRole.ADMIN:
#         return jsonify({'success': False, 'message': 'غير مسموح لك بالقيام بهذا الإجراء'})

#     operation = OperationRequest.query.get_or_404(operation_id)
#     review_notes = request.form.get('review_notes', '').strip()

#     if not review_notes:
#         return jsonify({'success': False, 'message': 'يجب إدخال سبب الرفض'})

#     try:
#         # 1. تحديث حالة الطلب
#         operation.status = 'rejected'
#         operation.reviewed_by = current_user.id
#         operation.reviewed_at = datetime.utcnow()
#         operation.review_notes = review_notes

#         # 2. البحث عن السجل المرتبط وحذفه (إذا كان handover)
#         record_to_delete = None
#         if operation.operation_type == 'handover':
#             record_to_delete = VehicleHandover.query.get(operation.related_record_id)

#         if record_to_delete:
#             db.session.delete(record_to_delete)
#             log_audit(
#                 'delete_on_rejection', 'VehicleHandover', record_to_delete.id,
#                 f"تم حذف سجل Handover رقم {record_to_delete.id} تلقائياً بسبب رفض الطلب #{operation.id}"
#             )
#         else:
#             current_app.logger.warning(f"لم يتم العثور على سجل Handover رقم {operation.related_record_id} لحذفه بعد الرفض.")

#         # 3. إنشاء إشعار
#         create_notification(
#             operation_id=operation.id, user_id=operation.requested_by, notification_type='status_change',
#             title=f'❌ تم رفض طلبك: {operation.title}',
#             message=f'تم رفض طلبك من قبل {current_user.username}. السبب: {review_notes}'
#         )

#         db.session.commit()

#         log_audit('reject', 'operation_request', operation.id, f'تم رفض العملية: {operation.title}')
#         flash('تم رفض العملية وحذف السجل المؤقت.', 'warning')
#         return jsonify({'success': True, 'message': 'تم رفض العملية'})

#     except Exception as e:
#         db.session.rollback()
#         current_app.logger.error(f"خطأ في رفض العملية #{operation_id}: {str(e)}")
#         return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})


@operations_bp.route('/<int:operation_id>/reject', methods=['POST'])
@login_required
def reject_operation(operation_id):
    """رفض العملية"""
    
    if current_user.role != UserRole.ADMIN:
        return jsonify({'success': False, 'message': 'غير مسموح'})
    
    operation = OperationRequest.query.get_or_404(operation_id)
    review_notes = request.form.get('review_notes', '').strip()
    
    if not review_notes:
        return jsonify({'success': False, 'message': 'يجب إدخال سبب الرفض'})
    
    try:
        # تحديث حالة العملية
        operation.status = 'rejected'
        operation.reviewed_by = current_user.id
        operation.reviewed_at = datetime.utcnow()
        operation.review_notes = review_notes
        
        # إنشاء إشعار للطالب
        create_notification(
            operation_id=operation.id,
            user_id=operation.requested_by,
            notification_type='status_change',
            title=f'تم رفض العملية: {operation.title}',
            message=f'تم رفض العملية من قبل {current_user.username}.\nالسبب: {review_notes}'
        )


        record_to_delete = None
        if operation.operation_type == 'handover':
            record_to_delete = VehicleHandover.query.get(operation.related_record_id)

        if record_to_delete:
            db.session.delete(record_to_delete)
            log_audit(
                'delete_on_rejection', 'VehicleHandover', record_to_delete.id,
                f"تم حذف سجل Handover رقم {record_to_delete.id} تلقائياً بسبب رفض الطلب #{operation.id}"
            )
        else:
            current_app.logger.warning(f"لم يتم العثور على سجل Handover رقم {operation.related_record_id} لحذفه بعد الرفض.")

        
        db.session.commit()
        
        # تسجيل العملية
        log_audit(current_user.id, 'reject', 'operation_request', operation.id, 
                 f'تم رفض العملية: {operation.title} - السبب: {review_notes}')
        
        flash('تم رفض العملية', 'warning')
        return jsonify({'success': True, 'message': 'تم رفض العملية'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})







@operations_bp.route('/<int:operation_id>/under-review', methods=['POST'])
@login_required  
def set_under_review(operation_id):
    """وضع العملية تحت المراجعة"""
    
    if current_user.role != UserRole.ADMIN:
        return jsonify({'success': False, 'message': 'غير مسموح'})
    
    operation = OperationRequest.query.get_or_404(operation_id)
    
    try:
        operation.status = 'under_review'
        operation.reviewed_by = current_user.id
        operation.reviewed_at = datetime.utcnow()
        
        # إنشاء إشعار للطالب
        create_notification(
            operation_id=operation.id,
            user_id=operation.requested_by,
            notification_type='status_change',
            title=f'العملية تحت المراجعة: {operation.title}',
            message=f'العملية قيد المراجعة من قبل {current_user.username}'
        )
        
        db.session.commit()
        
        log_audit(current_user.id, 'review', 'operation_request', operation.id, 
                 f'العملية تحت المراجعة: {operation.title}')
        
        return jsonify({'success': True, 'message': 'تم وضع العملية تحت المراجعة'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})



@operations_bp.route('/<int:id>/delete', methods=['GET', 'POST'])
@login_required
def delete_operation(id):
    """حذف العملية"""
    
    if current_user.role != UserRole.ADMIN:
        flash('غير مسموح لك بحذف العمليات', 'error')
        return redirect(url_for('operations.operations_list'))
    
    operation = OperationRequest.query.get_or_404(id)
    
    if request.method == 'GET':
        # عرض صفحة تأكيد الحذف
        return render_template('operations/delete.html', operation=operation)
    
    # معالجة الحذف عند POST
    try:
        # تسجيل العملية قبل الحذف
        operation_title = operation.title
        operation_type = operation.operation_type
        
        # حذف الإشعارات المرتبطة بالعملية أولاً
        OperationNotification.query.filter_by(operation_request_id=id).delete()
        
        # حذف العملية
        db.session.delete(operation)
        db.session.commit()
        
        # تسجيل عملية الحذف
        log_audit(current_user.id, 'delete', 'operation_request', id, f'تم حذف العملية: {operation_title} من النوع {operation_type}')
        
        flash('تم حذف العملية بنجاح', 'success')
        return redirect(url_for('operations.operations_list'))
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"خطأ في حذف العملية #{id}: {str(e)}")
        flash(f'حدث خطأ أثناء حذف العملية: {str(e)}', 'error')
        return redirect(url_for('operations.view_operation', operation_id=id))

@operations_bp.route('/notifications')
@login_required
def notifications():
    """عرض الإشعارات"""
    
    user_notifications = OperationNotification.query.filter_by(
        user_id=current_user.id
    ).order_by(OperationNotification.created_at.desc()).all()
    
    # تحديد الإشعارات كمقروءة
    OperationNotification.query.filter_by(
        user_id=current_user.id, 
        is_read=False
    ).update({'is_read': True, 'read_at': datetime.utcnow()})
    
    db.session.commit()
    
    return render_template('operations/notifications.html', 
                         notifications=user_notifications)

def create_operation_request(operation_type, related_record_id, vehicle_id, 
                           title, description, requested_by, priority='normal'):
    """إنشاء طلب عملية جديد"""
    
    operation = OperationRequest()
    operation.operation_type = operation_type
    operation.related_record_id = related_record_id
    operation.vehicle_id = vehicle_id
    operation.title = title
    operation.description = description
    operation.requested_by = requested_by
    operation.requested_at = datetime.utcnow()
    operation.priority = priority
    operation.status = 'pending'
    
    try:
        db.session.add(operation)
        db.session.flush()  # للحصول على ID
        
        # إنشاء إشعارات للمديرين
        admins = User.query.filter_by(role=UserRole.ADMIN).all()
        for admin in admins:
            create_notification(
                operation_id=operation.id,
                user_id=admin.id,
                notification_type='new_operation',
                title=f'عملية جديدة تحتاج موافقة: {title}',
                message=f'عملية جديدة من نوع {get_operation_type_name(operation_type)} تحتاج للمراجعة والموافقة.'
            )
        
        # لا نحفظ هنا، الدالة المستدعية مسؤولة عن الحفظ
        return operation
    except Exception as e:
        print(f"خطأ في create_operation_request: {str(e)}")
        import traceback
        traceback.print_exc()
        raise e

def create_notification(operation_id, user_id, notification_type, title, message):
    """إنشاء إشعار جديد"""
    
    notification = OperationNotification()
    notification.operation_request_id = operation_id
    notification.user_id = user_id
    notification.notification_type = notification_type
    notification.title = title
    notification.message = message
    
    db.session.add(notification)
    return notification

def get_operation_type_name(operation_type):
    """الحصول على اسم نوع العملية بالعربية"""
    
    type_names = {
        'handover': 'تسليم/استلام مركبة',
        'workshop': 'عملية ورشة',
        'workshop_record': 'سجل ورشة',
        'external_authorization': 'تفويض خارجي',
        'safety_inspection': 'فحص سلامة'
    }
    
    return type_names.get(operation_type, operation_type)

# دالة مساعدة للحصول على عدد العمليات المعلقة للمدير
def get_pending_operations_count():
    """الحصول على عدد العمليات المعلقة"""
    return OperationRequest.query.filter_by(status='pending').count()

# دالة مساعدة للحصول على عدد الإشعارات غير المقروءة
def get_unread_notifications_count(user_id):
    """الحصول على عدد الإشعارات غير المقروءة للمستخدم"""
    return OperationNotification.query.filter_by(
        user_id=user_id, 
        is_read=False
    ).count()

@operations_bp.route('/api/count')
@login_required
def api_operations_count():
    """API للحصول على إحصائيات العمليات"""
    
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'غير مسموح'})
    
    pending_count = OperationRequest.query.filter_by(status='pending').count()
    under_review_count = OperationRequest.query.filter_by(status='under_review').count()
    
    return jsonify({
        'pending': pending_count,
        'under_review': under_review_count,
        'unread_notifications': get_unread_notifications_count(current_user.id)
    })

@operations_bp.route('/<int:operation_id>/export-excel')
@login_required
def export_operation_excel(operation_id):
    """تصدير جميع بيانات العملية إلى Excel"""
    
    if current_user.role != UserRole.ADMIN:
        flash('غير مسموح لك بالوصول لهذه الصفحة', 'danger')
        return redirect(url_for('dashboard.index'))
    
    try:
        # جلب العملية والبيانات المرتبطة
        operation = OperationRequest.query.get_or_404(operation_id)
        related_record = operation.get_related_record()
        
        # إنشاء ملف Excel مع تنسيق محسن
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        
        # شيت 1: معلومات العملية الأساسية
        ws1 = wb.active
        ws1.title = 'معلومات العملية'
        
        # تنسيقات الألوان والخطوط
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2D5AA0', end_color='2D5AA0', fill_type='solid')
        data_font = Font(name='Arial', size=11)
        alignment = Alignment(horizontal='right', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # رؤوس الأعمدة
        operation_headers = ['رقم العملية', 'نوع العملية', 'العنوان', 'الوصف', 'الحالة', 
                           'الأولوية', 'تاريخ الطلب', 'تاريخ المراجعة', 'طالب العملية', 
                           'مراجع العملية', 'ملاحظات المراجعة']
        
        for col_num, header in enumerate(operation_headers, 1):
            cell = ws1.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border
            ws1.column_dimensions[cell.column_letter].width = 15
        
        # البيانات
        operation_values = [
            operation.id,
            get_operation_type_name(operation.operation_type),
            operation.title or '',
            operation.description or '',
            get_status_name(operation.status),
            get_priority_name(operation.priority),
            operation.requested_at.strftime('%Y-%m-%d %H:%M:%S') if operation.requested_at else '',
            operation.reviewed_at.strftime('%Y-%m-%d %H:%M:%S') if operation.reviewed_at else '',
            operation.requester.name if operation.requester else '',
            operation.reviewer.name if operation.reviewer else '',
            operation.review_notes or ''
        ]
        
        for col_num, value in enumerate(operation_values, 1):
            cell = ws1.cell(row=2, column=col_num, value=value)
            cell.font = data_font
            cell.alignment = alignment
            cell.border = border
            
        # شيت 2: بيانات المركبة
        if operation.vehicle:
            vehicle = operation.vehicle
            ws2 = wb.create_sheet('بيانات المركبة')
            
            vehicle_headers = ['رقم اللوحة', 'نوع المركبة', 'الماركة', 'الموديل', 'السنة', 
                              'اللون', 'الحالة', 'ملاحظات']
            
            for col_num, header in enumerate(vehicle_headers, 1):
                cell = ws2.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = border
                ws2.column_dimensions[cell.column_letter].width = 18
            
            vehicle_values = [
                vehicle.plate_number or '',
                getattr(vehicle, 'type_of_car', '') or '',
                getattr(vehicle, 'make', '') or '',
                vehicle.model or '',
                str(vehicle.year) if vehicle.year else '',
                vehicle.color or '',
                vehicle.status or '',
                vehicle.notes or ''
            ]
            
            for col_num, value in enumerate(vehicle_values, 1):
                cell = ws2.cell(row=2, column=col_num, value=value)
                cell.font = data_font
                cell.alignment = alignment
                cell.border = border
            
        # شيت 3: بيانات السائق/الموظف
        employee = None
        current_driver_info = None
        
        # البحث عن السائق من العملية الحالية أولاً
        if operation.operation_type == 'handover' and related_record:
            current_driver_info = {
                'name': getattr(related_record, 'person_name', ''),
                'phone': getattr(related_record, 'driver_phone_number', ''),
                'residency_number': getattr(related_record, 'driver_residency_number', '')
            }
            
            # البحث عن الموظف في النظام بالهوية
            if related_record.driver_residency_number:
                employee = Employee.query.filter_by(
                    national_id=related_record.driver_residency_number
                ).first()
            
            # البحث بالاسم إذا لم نجد بالهوية
            if not employee and related_record.person_name:
                employee = Employee.query.filter_by(
                    name=related_record.person_name
                ).first()
        
        # إذا لم نجد من العملية الحالية، ابحث من آخر تسليم للمركبة
        if not employee and not current_driver_info and operation.vehicle_id:
            last_handover = VehicleHandover.query.filter_by(
                vehicle_id=operation.vehicle_id,
                handover_type='delivery'
            ).order_by(VehicleHandover.handover_date.desc()).first()
            
            if last_handover:
                current_driver_info = {
                    'name': getattr(last_handover, 'person_name', ''),
                    'phone': getattr(last_handover, 'driver_phone_number', ''),
                    'residency_number': getattr(last_handover, 'driver_residency_number', '')
                }
                
                # البحث عن الموظف في النظام
                if last_handover.driver_residency_number:
                    employee = Employee.query.filter_by(
                        national_id=last_handover.driver_residency_number
                    ).first()
                
                if not employee and last_handover.person_name:
                    employee = Employee.query.filter_by(
                        name=last_handover.person_name
                    ).first()
        
        if employee or current_driver_info:
            ws3 = wb.create_sheet('بيانات السائق')
            
            driver_headers = ['الاسم', 'الرقم الوظيفي', 'رقم الهوية', 'رقم الجوال', 'رقم جوال العمل',
                             'رقم IMEI', 'القسم', 'المنصب', 'تاريخ التوظيف', 'الحالة', 
                             'البريد الإلكتروني', 'العنوان', 'تاريخ الميلاد', 'الجنسية']
            
            for col_num, header in enumerate(driver_headers, 1):
                cell = ws3.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = border
                ws3.column_dimensions[cell.column_letter].width = 16
            
            # ملء البيانات - أولوية للموظف المسجل في النظام
            if employee:
                driver_values = [
                    employee.name or '',
                    getattr(employee, 'employee_id', '') or '',
                    employee.national_id or '',
                    employee.mobile or operation.requester.phone or '966501234567',
                    getattr(employee, 'mobilePersonal', '') or '966507654321',
                    getattr(employee, 'mobile_imei', '') or '123456789012345',
                    employee.departments[0].name if employee.departments else '',
                    employee.job_title or '',
                    employee.join_date.strftime('%Y-%m-%d') if employee.join_date else '',
                    employee.status or '',
                    employee.email or '',
                    getattr(employee, 'location', '') or '',
                    employee.birth_date.strftime('%Y-%m-%d') if employee.birth_date else '',
                    employee.nationality or ''
                ]
            elif current_driver_info:
                # استخدام بيانات السائق من نموذج التسليم مع القيم الافتراضية
                driver_values = [
                    current_driver_info.get('name', '') or operation.requester.username or '',
                    '', # الرقم الوظيفي - غير متوفر
                    current_driver_info.get('residency_number', '') or '',
                    current_driver_info.get('phone', '') or operation.requester.phone or '966501234567',
                    '966507654321', # رقم جوال العمل
                    '123456789012345', # رقم IMEI
                    '', # القسم - غير متوفر
                    '', # المنصب - غير متوفر
                    '', # تاريخ التوظيف - غير متوفر
                    '', # الحالة - غير متوفر
                    operation.requester.email or '', # البريد الإلكتروني
                    '', # العنوان - غير متوفر
                    '', # تاريخ الميلاد - غير متوفر
                    ''  # الجنسية - غير متوفر
                ]
            else:
                # لا توجد بيانات سائق - استخدام بيانات المستخدم الطالب
                driver_values = [
                    operation.requester.username or '',
                    '', # الرقم الوظيفي - غير متوفر
                    '', # رقم الهوية
                    operation.requester.phone or '966501234567',
                    '966507654321', # رقم جوال العمل
                    '123456789012345', # رقم IMEI
                    '', # القسم - غير متوفر
                    '', # المنصب - غير متوفر
                    '', # تاريخ التوظيف - غير متوفر
                    'نشط', # الحالة
                    operation.requester.email or '', # البريد الإلكتروني
                    '', # العنوان - غير متوفر
                    '', # تاريخ الميلاد - غير متوفر
                    ''  # الجنسية - غير متوفر
                ]
            
            # كتابة البيانات في الخلايا
            for col_num, value in enumerate(driver_values, 1):
                cell = ws3.cell(row=2, column=col_num, value=value)
                cell.font = data_font
                cell.alignment = alignment
                cell.border = border
            
        # شيت 4: تفاصيل نموذج التسليم/الاستلام
        if related_record and hasattr(related_record, 'handover_type'):
            ws4 = wb.create_sheet('نموذج التسليم-الاستلام')
            
            handover_headers = ['نوع العملية', 'تاريخ العملية', 'اسم الشخص', 'رقم الجوال', 
                               'رقم الهوية', 'قراءة العداد', 'مستوى الوقود', 'حالة المركبة',
                               'ملاحظات', 'الموقع', 'تاريخ الإنشاء', 'أنشئ بواسطة', 'مصدر الإنشاء', 'رابط النموذج']
            
            for col_num, header in enumerate(handover_headers, 1):
                cell = ws4.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = border
                ws4.column_dimensions[cell.column_letter].width = 15
            
            # إنشاء رابط النموذج
            from flask import request
            base_url = request.host_url.rstrip('/')
            pdf_link = f"{base_url}/vehicles/handover/{related_record.id}/pdf/public"
            
            handover_values = [
                'تسليم' if related_record.handover_type == 'delivery' else 'استلام',
                related_record.handover_date.strftime('%Y-%m-%d %H:%M:%S') if related_record.handover_date else '',
                getattr(related_record, 'person_name', '') or '',
                getattr(related_record, 'driver_phone_number', '') or '',
                getattr(related_record, 'driver_residency_number', '') or '',
                str(getattr(related_record, 'mileage', '')) if getattr(related_record, 'mileage', None) else '',
                getattr(related_record, 'fuel_level', '') or '',
                getattr(related_record, 'vehicle_condition', '') or '',
                getattr(related_record, 'notes', '') or '',
                getattr(related_record, 'location', '') or '',
                related_record.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(related_record, 'created_at') and related_record.created_at else '',
                getattr(related_record, 'created_by_user', None).name if getattr(related_record, 'created_by_user', None) else '',
                'موبايل' if getattr(related_record, 'created_via_mobile', False) else 'ويب',
                pdf_link
            ]
            
            for col_num, value in enumerate(handover_values, 1):
                cell = ws4.cell(row=2, column=col_num, value=value)
                cell.font = data_font
                cell.alignment = alignment
                cell.border = border
            
        # شيت 5: سجلات الورشة (إذا كانت متوفرة)
        if operation.operation_type == 'workshop_record' and related_record:
            ws5 = wb.create_sheet('سجل الورشة')
            
            workshop_headers = ['نوع الخدمة', 'وصف المشكلة', 'الحل المطبق', 'التكلفة',
                               'تاريخ الدخول', 'تاريخ الخروج', 'الحالة', 'الفني المسؤول', 'ملاحظات']
            
            for col_num, header in enumerate(workshop_headers, 1):
                cell = ws5.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = border
                ws5.column_dimensions[cell.column_letter].width = 16
            
            workshop_values = [
                getattr(related_record, 'service_type', '') or '',
                getattr(related_record, 'problem_description', '') or '',
                getattr(related_record, 'solution_applied', '') or '',
                str(getattr(related_record, 'cost', '')) if getattr(related_record, 'cost', None) else '',
                related_record.entry_date.strftime('%Y-%m-%d') if getattr(related_record, 'entry_date', None) else '',
                related_record.exit_date.strftime('%Y-%m-%d') if getattr(related_record, 'exit_date', None) else '',
                getattr(related_record, 'status', '') or '',
                getattr(related_record, 'technician_name', '') or '',
                getattr(related_record, 'notes', '') or ''
            ]
            
            for col_num, value in enumerate(workshop_values, 1):
                cell = ws5.cell(row=2, column=col_num, value=value)
                cell.font = data_font
                cell.alignment = alignment
                cell.border = border
        
        # حفظ الملف
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # إنشاء اسم الملف
        filename = f"operation_details_{operation.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # تسجيل العملية
        log_audit(
            user_id=current_user.id,
            action='export',
            entity_type='operation_request',
            details=f'تصدير تفاصيل العملية {operation_id} إلى Excel'
        )
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'خطأ في تصدير البيانات: {str(e)}', 'danger')
        return redirect(url_for('operations.view_operation', operation_id=operation_id))

def get_status_name(status):
    """تحويل حالة العملية إلى النص العربي"""
    status_names = {
        'pending': 'معلقة',
        'under_review': 'تحت المراجعة',
        'approved': 'موافق عليها',
        'rejected': 'مرفوضة'
    }
    return status_names.get(status, status)

def get_priority_name(priority):
    """تحويل أولوية العملية إلى النص العربي"""
    priority_names = {
        'urgent': 'عاجل',
        'high': 'عالي',
        'normal': 'عادي',
        'low': 'منخفض'
    }
    return priority_names.get(priority, priority)