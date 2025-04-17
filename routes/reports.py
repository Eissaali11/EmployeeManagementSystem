from flask import Blueprint, render_template, request, jsonify, make_response, send_file
from sqlalchemy import func
from datetime import datetime, date, timedelta
from io import BytesIO
from utils.pdf import create_pdf, arabic_text, create_data_table, get_styles
from app import db
from models import Department, Employee, Attendance, Salary, Document, SystemAudit
from utils.date_converter import parse_date, format_date_hijri, format_date_gregorian, get_month_name_ar
from utils.excel import generate_employee_excel, generate_salary_excel
from utils.pdf_generator import generate_salary_report_pdf
# إضافة الاستيرادات المفقودة
import arabic_reshaper
from bidi.algorithm import get_display
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')

@reports_bp.route('/')
def index():
    """الصفحة الرئيسية للتقارير"""
    departments = Department.query.all()
    return render_template('reports/index.html', departments=departments)

@reports_bp.route('/employees')
def employees_report():
    """تقرير الموظفين حسب القسم"""
    department_id = request.args.get('department_id', '')
    status = request.args.get('status', '')
    
    query = Employee.query
    
    # تطبيق الفلاتر
    if department_id:
        query = query.filter_by(department_id=department_id)
    if status:
        query = query.filter_by(status=status)
    
    employees = query.all()
    departments = Department.query.all()
    
    return render_template('reports/employees.html', 
                          employees=employees, 
                          departments=departments,
                          department_id=department_id,
                          status=status)

@reports_bp.route('/employees/pdf')
def employees_pdf():
    """تصدير تقرير الموظفين إلى PDF"""
    department_id = request.args.get('department_id', '')
    status = request.args.get('status', '')
    
    query = Employee.query
    
    # تطبيق الفلاتر
    if department_id:
        query = query.filter_by(department_id=department_id)
        department = Department.query.get(department_id)
        department_name = department.name if department else ""
    else:
        department_name = "جميع الأقسام"
    
    if status:
        query = query.filter_by(status=status)
        if status == 'active':
            status_name = "نشط"
        elif status == 'inactive':
            status_name = "غير نشط"
        elif status == 'on_leave':
            status_name = "في إجازة"
        else:
            status_name = ""
    else:
        status_name = "جميع الحالات"
    
    employees = query.all()
    
    # استخدام المكتبة الموحدة لإنشاء PDF
    from utils.pdf import arabic_text, create_pdf, create_data_table, get_styles
    from reportlab.lib.units import cm
    from reportlab.platypus import Spacer, Paragraph
    
    # تجهيز العناصر
    elements = []
    
    # إضافة العنوان
    styles = get_styles()
    title = f"تقرير الموظفين - {department_name} - {status_name}"
    elements.append(Paragraph(arabic_text(title), styles['title']))
    elements.append(Spacer(1, 20))
    
    # إضافة تاريخ التقرير
    date_text = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d')}"
    elements.append(Paragraph(arabic_text(date_text), styles['arabic']))
    elements.append(Spacer(1, 20))
    
    # إعداد جدول البيانات
    headers = ["الاسم", "الرقم الوظيفي", "الرقم الوطني", "الهاتف", "المسمى الوظيفي", "القسم", "الحالة"]
    data = []
    
    # إضافة بيانات الموظفين
    for emp in employees:
        department_name = emp.department.name if emp.department else "---"
        
        # ترجمة حالة الموظف
        status_map = {
            'active': 'نشط',
            'inactive': 'غير نشط',
            'on_leave': 'في إجازة'
        }
        status_text = status_map.get(emp.status, emp.status)
        
        row = [
            arabic_text(emp.name),
            emp.employee_id,
            emp.national_id,
            emp.mobile,
            arabic_text(emp.job_title),
            arabic_text(department_name),
            arabic_text(status_text)
        ]
        data.append(row)
    
    # إنشاء الجدول
    if data:
        col_widths = [3*cm, 2*cm, 2*cm, 2*cm, 3*cm, 3*cm, 2*cm]
        table = create_data_table(headers, data, col_widths)
        elements.append(table)
    else:
        no_data_text = "لا توجد بيانات متاحة"
        elements.append(Paragraph(arabic_text(no_data_text), styles['arabic']))
    
    # إنشاء ملف PDF
    buffer = create_pdf(elements, landscape_mode=True)
    
    # إنشاء استجابة تحميل
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=employees_report.pdf'
    
    return response

@reports_bp.route('/employees/excel')
def employees_excel():
    """تصدير تقرير الموظفين إلى Excel"""
    department_id = request.args.get('department_id', '')
    status = request.args.get('status', '')
    
    query = Employee.query
    
    # تطبيق الفلاتر
    if department_id:
        query = query.filter_by(department_id=department_id)
    if status:
        query = query.filter_by(status=status)
    
    employees = query.all()
    
    # توليد ملف Excel
    output = generate_employee_excel(employees)
    
    # إنشاء استجابة تحميل
    return send_file(
        output,
        as_attachment=True,
        download_name=f'employees_report_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@reports_bp.route('/attendance')
def attendance_report():
    """تقرير الحضور والغياب"""
    # الحصول على معلمات الفلتر
    from_date_str = request.args.get('from_date', (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d'))
    to_date_str = request.args.get('to_date', datetime.now().strftime('%Y-%m-%d'))
    department_id = request.args.get('department_id', '')
    status = request.args.get('status', '')
    
    # معالجة التواريخ
    try:
        from_date = parse_date(from_date_str)
        to_date = parse_date(to_date_str)
    except ValueError:
        from_date = datetime.now() - timedelta(days=7)
        to_date = datetime.now()
    
    # استعلام الحضور
    query = db.session.query(
            Attendance, Employee
        ).join(
            Employee, Attendance.employee_id == Employee.id
        ).filter(
            Attendance.date.between(from_date, to_date)
        )
    
    # تطبيق الفلاتر
    if department_id:
        query = query.filter(Employee.department_id == department_id)
    if status:
        query = query.filter(Attendance.status == status)
    
    # الحصول على النتائج النهائية
    results = query.order_by(Attendance.date.desc()).all()
    
    # الحصول على قائمة الأقسام لعناصر الفلتر
    departments = Department.query.all()
    
    return render_template('reports/attendance.html',
                        results=results,
                        departments=departments,
                        from_date=from_date,
                        to_date=to_date,
                        department_id=department_id,
                        status=status,
                        format_date_gregorian=format_date_gregorian,
                        format_date_hijri=format_date_hijri)

@reports_bp.route('/attendance/pdf')
def attendance_pdf():
    """تصدير تقرير الحضور إلى PDF"""
    # الحصول على معلمات الفلتر
    from_date_str = request.args.get('from_date', (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d'))
    to_date_str = request.args.get('to_date', datetime.now().strftime('%Y-%m-%d'))
    department_id = request.args.get('department_id', '')
    status = request.args.get('status', '')
    
    # معالجة التواريخ
    try:
        from_date = parse_date(from_date_str)
        to_date = parse_date(to_date_str)
    except ValueError:
        from_date = datetime.now() - timedelta(days=7)
        to_date = datetime.now()
    
    # استعلام الحضور
    query = db.session.query(
            Attendance, Employee
        ).join(
            Employee, Attendance.employee_id == Employee.id
        ).filter(
            Attendance.date.between(from_date, to_date)
        )
    
    # تطبيق الفلاتر
    if department_id:
        query = query.filter(Employee.department_id == department_id)
        department = Department.query.get(department_id)
        department_name = department.name if department else ""
    else:
        department_name = "جميع الأقسام"
    
    if status:
        query = query.filter(Attendance.status == status)
        if status == 'present':
            status_name = "حاضر"
        elif status == 'absent':
            status_name = "غائب"
        elif status == 'leave':
            status_name = "إجازة"
        elif status == 'sick':
            status_name = "مرضي"
        else:
            status_name = ""
    else:
        status_name = "جميع الحالات"
    
    # الحصول على النتائج النهائية
    results = query.order_by(Attendance.date.desc()).all()
    
    # إنشاء ملف PDF
    buffer = BytesIO()
    
    # تسجيل الخط العربي
    try:
        # محاولة تسجيل الخط العربي إذا لم يكن مسجلاً مسبقًا
        pdfmetrics.registerFont(TTFont('Arabic', 'static/fonts/Arial.ttf'))
    except:
        # إذا كان هناك خطأ، نستخدم الخط الافتراضي
        pass
    
    # تعيين أبعاد الصفحة واتجاهها
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4),
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    # إعداد الأنماط
    styles = getSampleStyleSheet()
    # إنشاء نمط للنص العربي
    arabic_style = ParagraphStyle(
        name='Arabic',
        parent=styles['Normal'],
        fontName='Arabic',
        fontSize=12,
        alignment=1, # وسط
        textColor=colors.black
    )
    
    # إنشاء نمط للعناوين
    title_style = ParagraphStyle(
        name='Title',
        parent=styles['Title'],
        fontName='Arabic',
        fontSize=16,
        alignment=1, # وسط
        textColor=colors.black
    )
    
    # إعداد المحتوى
    elements = []
    
    # إضافة العنوان
    title = f"تقرير الحضور والغياب - {department_name} - {status_name}"
    # تهيئة النص العربي للعرض في PDF
    title = get_display(arabic_reshaper.reshape(title))
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 10))
    
    # إضافة نطاق التاريخ
    date_range = f"الفترة من: {format_date_gregorian(from_date)} إلى: {format_date_gregorian(to_date)}"
    date_range = get_display(arabic_reshaper.reshape(date_range))
    elements.append(Paragraph(date_range, arabic_style))
    elements.append(Spacer(1, 20))
    
    # إعداد جدول البيانات
    headers = ["التاريخ", "الاسم", "الرقم الوظيفي", "وقت الحضور", "وقت الانصراف", "الحالة", "القسم"]
    data = []
    
    # إضافة الرؤوس
    headers_display = [get_display(arabic_reshaper.reshape(h)) for h in headers]
    data.append(headers_display)
    
    # إضافة بيانات الحضور
    for attendance, employee in results:
        department_name = employee.department.name if employee.department else "---"
        
        # ترجمة حالة الحضور
        status_map = {
            'present': 'حاضر',
            'absent': 'غائب',
            'leave': 'إجازة',
            'sick': 'مرضي'
        }
        status_text = status_map.get(attendance.status, attendance.status)
        
        row = [
            format_date_gregorian(attendance.date),
            get_display(arabic_reshaper.reshape(employee.name)),
            employee.employee_id,
            str(attendance.check_in) if attendance.check_in else "---",
            str(attendance.check_out) if attendance.check_out else "---",
            get_display(arabic_reshaper.reshape(status_text)),
            get_display(arabic_reshaper.reshape(department_name))
        ]
        data.append(row)
    
    # إنشاء الجدول
    if data:
        # حساب العرض المناسب للجدول بناءً على حجم الصفحة
        table_width = landscape(A4)[0] - 4*cm  # العرض الإجمالي ناقص الهوامش
        col_widths = [table_width/len(headers)] * len(headers)  # توزيع متساوي
        table = Table(data, colWidths=col_widths)
        
        # إعداد أنماط الجدول
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),  # لون خلفية العناوين
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # لون نص العناوين
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # محاذاة النص
            ('FONTNAME', (0, 0), (-1, 0), 'Arabic'),  # خط العناوين
            ('FONTSIZE', (0, 0), (-1, 0), 12),  # حجم خط العناوين
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # تباعد أسفل العناوين
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # لون خلفية البيانات
            ('FONTNAME', (0, 1), (-1, -1), 'Arabic'),  # خط البيانات
            ('FONTSIZE', (0, 1), (-1, -1), 10),  # حجم خط البيانات
            ('GRID', (0, 0), (-1, -1), 1, colors.black),  # حدود الجدول
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # محاذاة النص عموديا
        ])
        
        # تطبيق التناوب في ألوان الصفوف لتحسين القراءة
        for i in range(1, len(data)):
            if i % 2 == 0:
                table_style.add('BACKGROUND', (0, i), (-1, i), colors.whitesmoke)
        
        table.setStyle(table_style)
        elements.append(table)
    else:
        no_data_text = get_display(arabic_reshaper.reshape("لا توجد بيانات متاحة"))
        elements.append(Paragraph(no_data_text, arabic_style))
    
    # إضافة معلومات التقرير في أسفل الصفحة
    elements.append(Spacer(1, 20))
    footer_text = f"تاريخ إنشاء التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    footer_text = get_display(arabic_reshaper.reshape(footer_text))
    elements.append(Paragraph(footer_text, arabic_style))
    
    # بناء المستند
    doc.build(elements)
    
    # إعادة المؤشر إلى بداية البايت
    buffer.seek(0)
    
    # إنشاء استجابة تحميل
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=attendance_report.pdf'
    
    return response

@reports_bp.route('/attendance/excel')
def attendance_excel():
    """تصدير تقرير الحضور إلى Excel"""
    # الحصول على معلمات الفلتر
    from_date_str = request.args.get('from_date', (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d'))
    to_date_str = request.args.get('to_date', datetime.now().strftime('%Y-%m-%d'))
    department_id = request.args.get('department_id', '')
    status = request.args.get('status', '')
    
    # معالجة التواريخ
    try:
        from_date = parse_date(from_date_str)
        to_date = parse_date(to_date_str)
    except ValueError:
        from_date = datetime.now() - timedelta(days=7)
        to_date = datetime.now()
    
    # استعلام الحضور
    query = db.session.query(
            Attendance, Employee
        ).join(
            Employee, Attendance.employee_id == Employee.id
        ).filter(
            Attendance.date.between(from_date, to_date)
        )
    
    # تطبيق الفلاتر
    if department_id:
        query = query.filter(Employee.department_id == department_id)
    if status:
        query = query.filter(Attendance.status == status)
    
    # الحصول على النتائج النهائية
    results = query.order_by(Attendance.date.desc()).all()
    
    # إنشاء كائن Pandas DataFrame
    import pandas as pd
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # إنشاء ملف Excel جديد
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "تقرير الحضور"
    
    # تنسيق العنوان
    sheet.merge_cells('A1:H1')
    title_cell = sheet['A1']
    title_cell.value = f"تقرير الحضور والغياب: {format_date_gregorian(from_date)} إلى {format_date_gregorian(to_date)}"
    title_cell.font = Font(size=16, bold=True, name='Calibri')
    title_cell.alignment = Alignment(horizontal='center')
    
    # إضافة عناوين الأعمدة
    headers = ["التاريخ", "الاسم", "الرقم الوظيفي", "وقت الحضور", "وقت الانصراف", "الحالة", "القسم", "ملاحظات"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True, name='Calibri')
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # تنسيق الحدود
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.border = thin_border
    
    # إضافة البيانات
    for idx, (attendance, employee) in enumerate(results, start=3):
        sheet.cell(row=idx, column=1).value = format_date_gregorian(attendance.date)
        sheet.cell(row=idx, column=2).value = employee.name
        sheet.cell(row=idx, column=3).value = employee.employee_id
        sheet.cell(row=idx, column=4).value = str(attendance.check_in) if attendance.check_in else "---"
        sheet.cell(row=idx, column=5).value = str(attendance.check_out) if attendance.check_out else "---"
        
        # ترجمة حالة الحضور
        status_ar = {
            'present': 'حاضر',
            'absent': 'غائب',
            'leave': 'إجازة',
            'sick': 'مرضي'
        }
        sheet.cell(row=idx, column=6).value = status_ar.get(attendance.status, attendance.status)
        
        # القسم
        department_name = employee.department.name if employee.department else "---"
        sheet.cell(row=idx, column=7).value = department_name
        
        # ملاحظات
        sheet.cell(row=idx, column=8).value = attendance.notes if attendance.notes else ""
        
        # تطبيق التنسيق على كل خلية
        for col in range(1, 9):
            cell = sheet.cell(row=idx, column=col)
            cell.alignment = Alignment(horizontal='center')
            
            # تنسيق الحدود
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
    
    # ضبط عرض الأعمدة
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # الحصول على حرف العمود
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width
    
    # ضبط اتجاه الورقة للعربية (من اليمين لليسار)
    sheet.sheet_view.rightToLeft = True
    
    # حفظ الملف
    workbook.save(output)
    output.seek(0)
    
    # إنشاء استجابة تحميل
    return send_file(
        output,
        as_attachment=True,
        download_name=f'attendance_report_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@reports_bp.route('/salaries')
def salaries_report():
    """تقرير الرواتب"""
    # الحصول على معلمات الفلتر
    current_year = datetime.now().year
    current_month = datetime.now().month
    
    month = int(request.args.get('month', current_month))
    year = int(request.args.get('year', current_year))
    department_id = request.args.get('department_id', '')
    
    # استعلام الرواتب
    salaries_query = Salary.query.filter_by(
        month=month,
        year=year
    )
    
    # الحصول على قائمة الموظفين الذين لديهم رواتب (لعرضها في التقرير)
    employee_ids_with_salary = [s.employee_id for s in salaries_query.all()]
    
    # استعلام لجميع الموظفين (لإظهار الموظفين الذين ليس لديهم رواتب مسجلة)
    employees_query = Employee.query.filter(Employee.status == 'active')
    
    if department_id:
        employees_query = employees_query.filter_by(department_id=department_id)
    
    employees = employees_query.all()
    
    # الحصول على الرواتب النهائية مع التفاصيل
    salaries = []
    for employee in employees:
        salary = Salary.query.filter_by(
            employee_id=employee.id,
            month=month,
            year=year
        ).first()
        
        if salary:
            salaries.append({
                'id': salary.id,
                'employee': employee,
                'basic_salary': salary.basic_salary,
                'allowances': salary.allowances,
                'deductions': salary.deductions,
                'bonus': salary.bonus,
                'net_salary': salary.net_salary,
                'has_salary': True
            })
        else:
            salaries.append({
                'id': None,
                'employee': employee,
                'basic_salary': 0,
                'allowances': 0,
                'deductions': 0,
                'bonus': 0,
                'net_salary': 0,
                'has_salary': False
            })
    
    # حساب الإجماليات
    totals = {
        'basic': sum(s['basic_salary'] for s in salaries if s['has_salary']),
        'allowances': sum(s['allowances'] for s in salaries if s['has_salary']),
        'deductions': sum(s['deductions'] for s in salaries if s['has_salary']),
        'bonus': sum(s['bonus'] for s in salaries if s['has_salary']),
        'net': sum(s['net_salary'] for s in salaries if s['has_salary'])
    }
    
    # الحصول على قائمة الأقسام لعناصر الفلتر
    departments = Department.query.all()
    
    return render_template('reports/salaries.html',
                        salaries=salaries,
                        departments=departments,
                        month=month,
                        year=year,
                        department_id=department_id,
                        totals=totals,
                        month_name=get_month_name_ar(month))

@reports_bp.route('/salaries/pdf')
def salaries_pdf():
    """تصدير تقرير الرواتب إلى PDF"""
    # الحصول على معلمات الفلتر
    current_year = datetime.now().year
    current_month = datetime.now().month
    
    month = int(request.args.get('month', current_month))
    year = int(request.args.get('year', current_year))
    department_id = request.args.get('department_id', '')
    
    # استعلام الرواتب
    salaries_query = Salary.query.filter_by(
        month=month,
        year=year
    )
    
    # الحصول على قائمة الموظفين الذين لديهم رواتب (لعرضها في التقرير)
    employee_ids_with_salary = [s.employee_id for s in salaries_query.all()]
    
    # استعلام لجميع الموظفين (لإظهار الموظفين الذين ليس لديهم رواتب مسجلة)
    employees_query = Employee.query.filter(Employee.status == 'active')
    
    if department_id:
        employees_query = employees_query.filter_by(department_id=department_id)
        department = Department.query.get(department_id)
        department_name = department.name if department else ""
    else:
        department_name = "جميع الأقسام"
    
    employees = employees_query.all()
    
    # الحصول على الرواتب النهائية مع التفاصيل
    salaries = []
    for employee in employees:
        salary = Salary.query.filter_by(
            employee_id=employee.id,
            month=month,
            year=year
        ).first()
        
        if salary:
            salaries.append({
                'id': salary.id,
                'employee': employee,
                'basic_salary': salary.basic_salary,
                'allowances': salary.allowances,
                'deductions': salary.deductions,
                'bonus': salary.bonus,
                'net_salary': salary.net_salary,
                'has_salary': True
            })
    
    # حساب الإجماليات
    totals = {
        'basic': sum(s['basic_salary'] for s in salaries if s['has_salary']),
        'allowances': sum(s['allowances'] for s in salaries if s['has_salary']),
        'deductions': sum(s['deductions'] for s in salaries if s['has_salary']),
        'bonus': sum(s['bonus'] for s in salaries if s['has_salary']),
        'net': sum(s['net_salary'] for s in salaries if s['has_salary'])
    }
    
    # إنشاء ملف PDF
    buffer = BytesIO()
    
    # تسجيل الخط العربي
    try:
        # محاولة تسجيل الخط العربي إذا لم يكن مسجلاً مسبقًا
        pdfmetrics.registerFont(TTFont('Arabic', 'static/fonts/Arial.ttf'))
    except:
        # إذا كان هناك خطأ، نستخدم الخط الافتراضي
        pass
    
    # تعيين أبعاد الصفحة واتجاهها
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4),
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    # إعداد الأنماط
    styles = getSampleStyleSheet()
    # إنشاء نمط للنص العربي
    arabic_style = ParagraphStyle(
        name='Arabic',
        parent=styles['Normal'],
        fontName='Arabic',
        fontSize=12,
        alignment=1, # وسط
        textColor=colors.black
    )
    
    # إنشاء نمط للعناوين
    title_style = ParagraphStyle(
        name='Title',
        parent=styles['Title'],
        fontName='Arabic',
        fontSize=16,
        alignment=1, # وسط
        textColor=colors.black
    )
    
    # إعداد المحتوى
    elements = []
    
    # إضافة العنوان
    title = f"كشف رواتب شهر {get_month_name_ar(month)} {year} - {department_name}"
    # تهيئة النص العربي للعرض في PDF
    title = get_display(arabic_reshaper.reshape(title))
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 20))
    
    # إعداد جدول البيانات
    headers = ["الاسم", "الرقم الوظيفي", "الراتب الأساسي", "البدلات", "الخصومات", "المكافآت", "صافي الراتب"]
    data = []
    
    # إضافة الرؤوس
    headers_display = [get_display(arabic_reshaper.reshape(h)) for h in headers]
    data.append(headers_display)
    
    # إضافة بيانات الرواتب
    for salary_item in salaries:
        if salary_item['has_salary']:
            employee = salary_item['employee']
            row = [
                get_display(arabic_reshaper.reshape(employee.name)),
                employee.employee_id,
                f"{salary_item['basic_salary']:.2f}",
                f"{salary_item['allowances']:.2f}",
                f"{salary_item['deductions']:.2f}",
                f"{salary_item['bonus']:.2f}",
                f"{salary_item['net_salary']:.2f}"
            ]
            data.append(row)
    
    # إنشاء الجدول
    if len(data) > 1:  # لدينا بيانات بخلاف الرؤوس
        # حساب العرض المناسب للجدول بناءً على حجم الصفحة
        table_width = landscape(A4)[0] - 4*cm  # العرض الإجمالي ناقص الهوامش
        col_widths = [table_width/len(headers)] * len(headers)  # توزيع متساوي
        table = Table(data, colWidths=col_widths)
        
        # إعداد أنماط الجدول
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),  # لون خلفية العناوين
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # لون نص العناوين
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # محاذاة النص
            ('FONTNAME', (0, 0), (-1, 0), 'Arabic'),  # خط العناوين
            ('FONTSIZE', (0, 0), (-1, 0), 12),  # حجم خط العناوين
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # تباعد أسفل العناوين
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # لون خلفية البيانات
            ('FONTNAME', (0, 1), (-1, -1), 'Arabic'),  # خط البيانات
            ('FONTSIZE', (0, 1), (-1, -1), 10),  # حجم خط البيانات
            ('GRID', (0, 0), (-1, -1), 1, colors.black),  # حدود الجدول
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # محاذاة النص عموديا
        ])
        
        # تطبيق التناوب في ألوان الصفوف لتحسين القراءة
        for i in range(1, len(data)):
            if i % 2 == 0:
                table_style.add('BACKGROUND', (0, i), (-1, i), colors.whitesmoke)
        
        table.setStyle(table_style)
        elements.append(table)
        
        # إضافة صف الإجماليات
        elements.append(Spacer(1, 20))
        totals_text = get_display(arabic_reshaper.reshape(f"الإجماليات: الراتب الأساسي: {totals['basic']:.2f} - البدلات: {totals['allowances']:.2f} - الخصومات: {totals['deductions']:.2f} - المكافآت: {totals['bonus']:.2f} - صافي الرواتب: {totals['net']:.2f}"))
        elements.append(Paragraph(totals_text, arabic_style))
    else:
        no_data_text = get_display(arabic_reshaper.reshape("لا توجد بيانات رواتب لهذه الفترة"))
        elements.append(Paragraph(no_data_text, arabic_style))
    
    # إضافة معلومات التقرير في أسفل الصفحة
    elements.append(Spacer(1, 20))
    footer_text = f"تاريخ إنشاء التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    footer_text = get_display(arabic_reshaper.reshape(footer_text))
    elements.append(Paragraph(footer_text, arabic_style))
    
    # بناء المستند
    doc.build(elements)
    
    # إعادة المؤشر إلى بداية البايت
    buffer.seek(0)
    
    # إنشاء استجابة تحميل
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=salaries_report.pdf'
    
    return response

@reports_bp.route('/salaries/excel')
def salaries_excel():
    """تصدير تقرير الرواتب إلى Excel"""
    # الحصول على معلمات الفلتر
    current_year = datetime.now().year
    current_month = datetime.now().month
    
    month = int(request.args.get('month', current_month))
    year = int(request.args.get('year', current_year))
    department_id = request.args.get('department_id', '')
    
    # استعلام الرواتب
    salaries_query = Salary.query.filter_by(
        month=month,
        year=year
    )
    
    # تطبيق الفلاتر
    if department_id:
        # نحتاج أولاً إلى الحصول على معرفات الموظفين في القسم المحدد
        dept_employee_ids = [e.id for e in Employee.query.filter_by(department_id=department_id).all()]
        salaries_query = salaries_query.filter(Salary.employee_id.in_(dept_employee_ids))
    
    salaries = salaries_query.all()
    
    # إنشاء كائن Pandas DataFrame
    import pandas as pd
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # إنشاء ملف Excel جديد
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "تقرير الرواتب"
    
    # تنسيق العنوان
    sheet.merge_cells('A1:G1')
    title_cell = sheet['A1']
    title_cell.value = f"تقرير الرواتب لشهر {get_month_name_ar(month)} {year}"
    title_cell.font = Font(size=16, bold=True, name='Calibri')
    title_cell.alignment = Alignment(horizontal='center')
    
    # إضافة عناوين الأعمدة
    headers = ["اسم الموظف", "الرقم الوظيفي", "القسم", "الراتب الأساسي", "البدلات", "الاستقطاعات", "المكافآت", "صافي الراتب"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True, name='Calibri')
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # تنسيق الحدود
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.border = thin_border
    
    # إضافة البيانات
    for idx, salary in enumerate(salaries, start=3):
        employee = salary.employee
        sheet.cell(row=idx, column=1).value = employee.name
        sheet.cell(row=idx, column=2).value = employee.employee_id
        
        # القسم
        department_name = employee.department.name if employee.department else "---"
        sheet.cell(row=idx, column=3).value = department_name
        
        # تفاصيل الراتب
        sheet.cell(row=idx, column=4).value = salary.basic_salary
        sheet.cell(row=idx, column=5).value = salary.allowances
        sheet.cell(row=idx, column=6).value = salary.deductions
        sheet.cell(row=idx, column=7).value = salary.bonus
        sheet.cell(row=idx, column=8).value = salary.net_salary
        
        # تطبيق التنسيق على كل خلية
        for col in range(1, 9):
            cell = sheet.cell(row=idx, column=col)
            cell.alignment = Alignment(horizontal='center')
            
            # تنسيق الحدود
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
    
    # إضافة صف الإجماليات
    total_row = len(salaries) + 3
    
    sheet.cell(row=total_row, column=1).value = "الإجمالي"
    sheet.cell(row=total_row, column=1).font = Font(bold=True, name='Calibri')
    sheet.merge_cells(f'A{total_row}:C{total_row}')
    
    # حساب الإجماليات
    basic_total = sum(s.basic_salary for s in salaries)
    allowances_total = sum(s.allowances for s in salaries)
    deductions_total = sum(s.deductions for s in salaries)
    bonus_total = sum(s.bonus for s in salaries)
    net_total = sum(s.net_salary for s in salaries)
    
    sheet.cell(row=total_row, column=4).value = basic_total
    sheet.cell(row=total_row, column=5).value = allowances_total
    sheet.cell(row=total_row, column=6).value = deductions_total
    sheet.cell(row=total_row, column=7).value = bonus_total
    sheet.cell(row=total_row, column=8).value = net_total
    
    # تنسيق صف الإجماليات
    for col in range(1, 9):
        cell = sheet.cell(row=total_row, column=col)
        cell.font = Font(bold=True, name='Calibri')
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
        
        # تنسيق الحدود
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
    
    # ضبط عرض الأعمدة
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # الحصول على حرف العمود
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width
    
    # ضبط اتجاه الورقة للعربية (من اليمين لليسار)
    sheet.sheet_view.rightToLeft = True
    
    # حفظ الملف
    workbook.save(output)
    output.seek(0)
    
    # إنشاء استجابة تحميل
    return send_file(
        output,
        as_attachment=True,
        download_name=f'salaries_report_{month}_{year}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@reports_bp.route('/documents')
def documents_report():
    """تقرير الوثائق"""
    # الحصول على معلمات الفلتر
    department_id = request.args.get('department_id', '')
    document_type = request.args.get('document_type', '')
    expiring_only = request.args.get('expiring_only', '') == 'true'
    expiry_days = int(request.args.get('expiry_days', 30))
    
    # تحديد تاريخ الانتهاء للمقارنة في حالة "قريبة من الانتهاء"
    cutoff_date = datetime.now().date() + timedelta(days=expiry_days)
    
    # استعلام الوثائق
    query = db.session.query(
            Document, Employee
        ).join(
            Employee, Document.employee_id == Employee.id
        )
    
    # تطبيق الفلاتر
    if department_id:
        query = query.filter(Employee.department_id == department_id)
    if document_type:
        query = query.filter(Document.document_type == document_type)
    if expiring_only:
        query = query.filter(Document.expiry_date <= cutoff_date)
    
    # الحصول على النتائج النهائية
    results = query.order_by(Document.expiry_date).all()
    
    # الحصول على قائمة الأقسام وأنواع الوثائق لعناصر الفلتر
    departments = Department.query.all()
    document_types = [
        {'id': 'national_id', 'name': 'الهوية الوطنية'},
        {'id': 'passport', 'name': 'جواز السفر'},
        {'id': 'driver_license', 'name': 'رخصة القيادة'},
        {'id': 'annual_leave', 'name': 'الإجازة السنوية'},
        {'id': 'health_certificate', 'name': 'الشهادة الصحية'}
    ]
    
    return render_template('reports/documents.html',
                        results=results,
                        departments=departments,
                        document_types=document_types,
                        department_id=department_id,
                        document_type=document_type,
                        expiring_only=expiring_only,
                        expiry_days=expiry_days,
                        format_date_gregorian=format_date_gregorian,
                        format_date_hijri=format_date_hijri)

@reports_bp.route('/documents/pdf')
def documents_pdf():
    """تصدير تقرير الوثائق إلى PDF"""
    # الحصول على معلمات الفلتر
    department_id = request.args.get('department_id', '')
    document_type = request.args.get('document_type', '')
    expiring_only = request.args.get('expiring_only', '') == 'true'
    expiry_days = int(request.args.get('expiry_days', 30))
    
    # تحديد تاريخ الانتهاء للمقارنة في حالة "قريبة من الانتهاء"
    cutoff_date = datetime.now().date() + timedelta(days=expiry_days)
    
    # استعلام الوثائق
    query = db.session.query(
            Document, Employee
        ).join(
            Employee, Document.employee_id == Employee.id
        )
    
    # تطبيق الفلاتر والحصول على أسماء الفلاتر
    if department_id:
        query = query.filter(Employee.department_id == department_id)
        department = Department.query.get(department_id)
        department_name = department.name if department else ""
    else:
        department_name = "جميع الأقسام"
    
    if document_type:
        query = query.filter(Document.document_type == document_type)
        document_types_map = {
            'national_id': 'الهوية الوطنية',
            'passport': 'جواز السفر',
            'driver_license': 'رخصة القيادة',
            'annual_leave': 'الإجازة السنوية',
            'health_certificate': 'الشهادة الصحية'
        }
        document_type_name = document_types_map.get(document_type, "")
    else:
        document_type_name = "جميع أنواع الوثائق"
    
    if expiring_only:
        query = query.filter(Document.expiry_date <= cutoff_date)
        expiry_status = f"الوثائق التي ستنتهي خلال {expiry_days} يوم"
    else:
        expiry_status = "جميع الوثائق"
    
    # الحصول على النتائج النهائية
    results = query.order_by(Document.expiry_date).all()
    
    # استخدام المكتبة الموحدة لإنشاء PDF
    from utils.pdf import arabic_text, create_pdf, create_data_table, get_styles
    from reportlab.lib.units import cm
    from reportlab.platypus import Spacer, Paragraph
    from reportlab.lib import colors
    
    # إعداد المحتوى باستخدام وحدة PDF الجديدة
    elements = []
    
    # الحصول على أنماط النصوص المحسنة
    styles = get_styles()
    
    # إضافة العنوان
    title = f"تقرير الوثائق - {document_type_name} - {department_name} - {expiry_status}"
    # تهيئة النص العربي للعرض في PDF باستخدام الدالة المحسنة
    elements.append(Paragraph(arabic_text(title), styles['title']))
    elements.append(Spacer(1, 20))
    
    # إعداد جدول البيانات
    headers = ["الموظف", "الرقم الوظيفي", "القسم", "نوع الوثيقة", "رقم الوثيقة", "تاريخ الإصدار", "تاريخ الانتهاء", "الحالة"]
    data = []
    
    # إضافة الرؤوس (مع تطبيق ترميز النص العربي باستخدام الدالة المحسنة)
    headers_display = [arabic_text(h) for h in headers]
    data.append(headers_display)
    
    # ترجمة أنواع الوثائق
    document_types_map = {
        'national_id': 'الهوية الوطنية',
        'passport': 'جواز السفر',
        'driver_license': 'رخصة القيادة',
        'annual_leave': 'الإجازة السنوية',
        'health_certificate': 'الشهادة الصحية'
    }
    
    # إضافة بيانات الوثائق
    today = datetime.now().date()
    for document, employee in results:
        department_name = employee.department.name if employee.department else "---"
        document_type_arabic = document_types_map.get(document.document_type, document.document_type)
        
        # تحديد حالة الوثيقة (سارية، قاربت الانتهاء، منتهية)
        days_to_expiry = (document.expiry_date - today).days
        if days_to_expiry <= 0:
            status = "منتهية"
            status_color = colors.red
        elif days_to_expiry <= expiry_days:
            status = f"تنتهي خلال {days_to_expiry} يوم"
            status_color = colors.orange
        else:
            status = "سارية"
            status_color = colors.green
        
        row = [
            arabic_text(employee.name),
            employee.employee_id,
            arabic_text(department_name),
            arabic_text(document_type_arabic),
            document.document_number,
            format_date_gregorian(document.issue_date),
            format_date_gregorian(document.expiry_date),
            arabic_text(status)
        ]
        data.append(row)
    
    # إنشاء الجدول باستخدام دالة جدول البيانات المحسنة
    if len(data) > 1:  # لدينا بيانات بخلاف الرؤوس
        # إنشاء جدول البيانات باستخدام الدالة المحسنة
        elements.append(create_data_table(headers, data[1:]))
    else:
        elements.append(Paragraph(arabic_text("لا توجد بيانات وثائق متاحة"), styles['normal']))
    
    # إضافة معلومات التقرير في أسفل الصفحة
    elements.append(Spacer(1, 20))
    footer_text = f"تاريخ إنشاء التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    elements.append(Paragraph(arabic_text(footer_text), styles['normal']))
    
    # إنشاء ملف PDF واستخدام دالة الإنشاء المحسنة
    buffer = create_pdf(elements, landscape_mode=True)
    
    # إنشاء استجابة تحميل
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=documents_report.pdf'
    
    return response

@reports_bp.route('/documents/excel')
def documents_excel():
    """تصدير تقرير الوثائق إلى Excel"""
    # الحصول على معلمات الفلتر
    department_id = request.args.get('department_id', '')
    document_type = request.args.get('document_type', '')
    expiring_only = request.args.get('expiring_only', '') == 'true'
    expiry_days = int(request.args.get('expiry_days', 30))
    
    # تحديد تاريخ الانتهاء للمقارنة في حالة "قريبة من الانتهاء"
    cutoff_date = datetime.now().date() + timedelta(days=expiry_days)
    
    # استعلام الوثائق
    query = db.session.query(
            Document, Employee
        ).join(
            Employee, Document.employee_id == Employee.id
        )
    
    # تطبيق الفلاتر
    if department_id:
        query = query.filter(Employee.department_id == department_id)
    if document_type:
        query = query.filter(Document.document_type == document_type)
    if expiring_only:
        query = query.filter(Document.expiry_date <= cutoff_date)
    
    # الحصول على النتائج النهائية
    results = query.order_by(Document.expiry_date).all()
    
    # إنشاء كائن Pandas DataFrame
    import pandas as pd
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # إنشاء ملف Excel جديد
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "تقرير الوثائق"
    
    # تحديد عنوان التقرير
    title = "تقرير الوثائق"
    if expiring_only:
        title += f" - الوثائق التي ستنتهي خلال {expiry_days} يوم"
    
    # تنسيق العنوان
    sheet.merge_cells('A1:H1')
    title_cell = sheet['A1']
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True, name='Calibri')
    title_cell.alignment = Alignment(horizontal='center')
    
    # إضافة عناوين الأعمدة
    headers = ["اسم الموظف", "الرقم الوظيفي", "القسم", "نوع الوثيقة", "رقم الوثيقة", "تاريخ الإصدار", "تاريخ الانتهاء", "الحالة", "ملاحظات"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True, name='Calibri')
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # تنسيق الحدود
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.border = thin_border
    
    # ترجمة أنواع الوثائق
    document_types_map = {
        'national_id': 'الهوية الوطنية',
        'passport': 'جواز السفر',
        'driver_license': 'رخصة القيادة',
        'annual_leave': 'الإجازة السنوية',
        'health_certificate': 'الشهادة الصحية'
    }
    
    # إضافة البيانات
    for idx, (document, employee) in enumerate(results, start=3):
        sheet.cell(row=idx, column=1).value = employee.name
        sheet.cell(row=idx, column=2).value = employee.employee_id
        
        # القسم
        department_name = employee.department.name if employee.department else "---"
        sheet.cell(row=idx, column=3).value = department_name
        
        # تفاصيل الوثيقة
        sheet.cell(row=idx, column=4).value = document_types_map.get(document.document_type, document.document_type)
        sheet.cell(row=idx, column=5).value = document.document_number
        sheet.cell(row=idx, column=6).value = format_date_gregorian(document.issue_date)
        sheet.cell(row=idx, column=7).value = format_date_gregorian(document.expiry_date)
        
        # حالة الوثيقة (سارية، منتهية، قريبة من الانتهاء)
        today = datetime.now().date()
        if document.expiry_date < today:
            status = "منتهية"
            status_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # أحمر فاتح
        elif document.expiry_date <= today + timedelta(days=30):
            status = "تنتهي قريباً"
            status_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # أصفر فاتح
        else:
            status = "سارية"
            status_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # أخضر فاتح
        
        status_cell = sheet.cell(row=idx, column=8)
        status_cell.value = status
        status_cell.fill = status_fill
        
        # ملاحظات
        sheet.cell(row=idx, column=9).value = document.notes if document.notes else ""
        
        # تطبيق التنسيق على كل خلية
        for col in range(1, 10):
            cell = sheet.cell(row=idx, column=col)
            cell.alignment = Alignment(horizontal='center')
            
            # تنسيق الحدود
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
    
    # ضبط عرض الأعمدة
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # الحصول على حرف العمود
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width
    
    # ضبط اتجاه الورقة للعربية (من اليمين لليسار)
    sheet.sheet_view.rightToLeft = True
    
    # حفظ الملف
    workbook.save(output)
    output.seek(0)
    
    # إنشاء استجابة تحميل
    return send_file(
        output,
        as_attachment=True,
        download_name=f'documents_report_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )