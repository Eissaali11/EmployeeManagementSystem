from openpyxl.utils import get_column_letter

def adjust_column_width_safely(sheet):
    """
    ضبط عرض الأعمدة بطريقة آمنة تتعامل مع الخلايا المدمجة
    """
    # تعيين العرض الافتراضي لجميع الأعمدة
    for i in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(i)].width = 15
    
    # نهاية الوظيفة - عدم تطبيق المزيد من ضبط العرض لتجنب أخطاء الخلايا المدمجة