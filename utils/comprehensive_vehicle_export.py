"""
تصدير شامل لجميع بيانات أقسام المركبة إلى Excel
"""
import os
from io import BytesIO
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def export_comprehensive_vehicle_excel(vehicle, workshop_records=None, rental_records=None, 
                                      project_records=None, handover_records=None, 
                                      inspection_records=None, safety_check_records=None, 
                                      accident_records=None):
    """
    تصدير شامل لجميع بيانات المركبة إلى ملف Excel متعدد الأوراق
    
    Args:
        vehicle: كائن المركبة
        workshop_records: سجلات الورشة
        rental_records: سجلات الإيجار
        project_records: سجلات المشاريع
        handover_records: سجلات التسليم والاستلام
        inspection_records: سجلات الفحص الدوري
        safety_check_records: سجلات فحوصات السلامة
        accident_records: سجلات الحوادث
    
    Returns:
        BytesIO: كائن بايت يحتوي على ملف Excel
    """
    
    buffer = BytesIO()
    
    try:
        # إنشاء مصنف Excel
        wb = openpyxl.Workbook()
        
        # إزالة الورقة الافتراضية
        if wb.active:
            wb.remove(wb.active)
        
        # إعداد الأنماط
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. ورقة المعلومات الأساسية
        ws_basic = wb.create_sheet("المعلومات الأساسية")
        
        # البيانات الأساسية
        basic_data = [
            ["البيان", "القيمة"],
            ["رقم اللوحة", vehicle.plate_number or ""],
            ["الماركة", vehicle.make or ""],
            ["الموديل", vehicle.model or ""],
            ["سنة الصنع", str(vehicle.year) if vehicle.year else ""],
            ["اللون", vehicle.color or ""],
            ["رقم الهيكل", vehicle.chassis_number or ""],
            ["رقم المحرك", vehicle.engine_number or ""],
            ["نوع الوقود", vehicle.fuel_type or ""],
            ["عداد المسافات", str(vehicle.mileage) if vehicle.mileage else ""],
            ["السائق الحالي", vehicle.driver_name or "غير محدد"],
            ["الحالة", {
                'available': 'متاحة',
                'rented': 'مؤجرة', 
                'in_workshop': 'في الورشة',
                'in_project': 'في المشروع',
                'accident': 'حادث',
                'sold': 'مباعة'
            }.get(vehicle.status, vehicle.status or "")],
            ["تاريخ انتهاء الفحص الدوري", vehicle.inspection_expiry_date.strftime("%Y-%m-%d") if vehicle.inspection_expiry_date else "غير محدد"],
            ["تاريخ انتهاء الاستمارة", vehicle.registration_expiry_date.strftime("%Y-%m-%d") if vehicle.registration_expiry_date else "غير محدد"],
            ["الملاحظات", vehicle.notes or ""]
        ]
        
        # كتابة البيانات الأساسية
        for row_idx, row_data in enumerate(basic_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_basic.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                if row_idx == 1:  # رأس الجدول
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                else:
                    if col_idx == 1:  # عمود البيان
                        cell.font = Font(bold=True)
                    cell.alignment = right_alignment
        
        # تعديل عرض الأعمدة
        ws_basic.column_dimensions['A'].width = 25
        ws_basic.column_dimensions['B'].width = 30
        
        # 2. ورقة سجلات الورشة
        if workshop_records and len(workshop_records) > 0:
            ws_workshop = wb.create_sheet("سجلات الورشة")
            
            workshop_headers = [
                "تاريخ الدخول", "تاريخ الخروج", "سبب الدخول", "حالة الإصلاح", 
                "التكلفة (ريال)", "اسم الورشة", "الفني المسؤول", "الوصف", "ملاحظات"
            ]
            
            # كتابة العناوين
            for col_idx, header in enumerate(workshop_headers, 1):
                cell = ws_workshop.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # كتابة البيانات
            reason_map = {'maintenance': 'صيانة دورية', 'breakdown': 'عطل', 'accident': 'حادث'}
            status_map = {'in_progress': 'قيد التنفيذ', 'completed': 'تم الإصلاح', 'pending_approval': 'بانتظار الموافقة'}
            
            total_cost = 0
            for row_idx, record in enumerate(workshop_records, 2):
                cost = record.cost or 0
                total_cost += cost
                
                row_data = [
                    record.entry_date.strftime("%Y-%m-%d") if record.entry_date else "غير محدد",
                    record.exit_date.strftime("%Y-%m-%d") if record.exit_date else "ما زالت في الورشة",
                    reason_map.get(record.reason, record.reason or "غير محدد"),
                    status_map.get(record.repair_status, record.repair_status or "غير محدد"),
                    cost,
                    record.workshop_name or "غير محدد",
                    record.technician_name or "غير محدد",
                    record.description or "",
                    record.notes or ""
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_workshop.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = right_alignment if col_idx != 5 else center_alignment
            
            # إضافة إجمالي التكلفة
            total_row = len(workshop_records) + 3
            ws_workshop.merge_cells(f'A{total_row}:D{total_row}')
            total_cell = ws_workshop.cell(row=total_row, column=1, value='الإجمالي')
            total_cell.font = header_font
            total_cell.fill = header_fill
            total_cell.alignment = center_alignment
            total_cell.border = thin_border
            
            cost_cell = ws_workshop.cell(row=total_row, column=5, value=total_cost)
            cost_cell.font = Font(bold=True)
            cost_cell.border = thin_border
            cost_cell.alignment = center_alignment
            
            # تعديل عرض الأعمدة
            column_widths = [15, 15, 20, 20, 12, 20, 20, 25, 25]
            for i, width in enumerate(column_widths, 1):
                from openpyxl.utils import get_column_letter
                ws_workshop.column_dimensions[get_column_letter(i)].width = width
        
        # 3. ورقة سجلات الإيجار
        if rental_records and len(rental_records) > 0:
            ws_rental = wb.create_sheet("سجلات الإيجار")
            
            rental_headers = [
                "المستأجر", "تاريخ البداية", "تاريخ النهاية", "التكلفة الشهرية (ريال)", 
                "الحالة", "جهة الاتصال", "رقم العقد", "المدينة", "ملاحظات"
            ]
            
            # كتابة العناوين
            for col_idx, header in enumerate(rental_headers, 1):
                cell = ws_rental.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # كتابة البيانات
            for row_idx, record in enumerate(rental_records, 2):
                row_data = [
                    record.lessor_name or "غير محدد",
                    record.start_date.strftime("%Y-%m-%d") if record.start_date else "غير محدد",
                    record.end_date.strftime("%Y-%m-%d") if record.end_date else "مستمر",
                    record.monthly_cost or 0,
                    "نشط" if record.is_active else "منتهي",
                    record.lessor_contact or "",
                    record.contract_number or "",
                    record.city or "",
                    record.notes or ""
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_rental.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = right_alignment if col_idx != 4 else center_alignment
            
            # تعديل عرض الأعمدة
            column_widths = [20, 15, 15, 15, 10, 15, 15, 15, 25]
            for i, width in enumerate(column_widths, 1):
                from openpyxl.utils import get_column_letter
                ws_rental.column_dimensions[get_column_letter(i)].width = width
        
        # 4. ورقة سجلات المشاريع
        if project_records and len(project_records) > 0:
            ws_projects = wb.create_sheet("سجلات المشاريع")
            
            project_headers = [
                "اسم المشروع", "الموقع", "تاريخ التكليف", "تاريخ الانتهاء", 
                "المدير المسؤول", "الحالة", "ملاحظات"
            ]
            
            # كتابة العناوين
            for col_idx, header in enumerate(project_headers, 1):
                cell = ws_projects.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # كتابة البيانات
            for row_idx, record in enumerate(project_records, 2):
                row_data = [
                    record.project_name or "غير محدد",
                    record.location or "غير محدد",
                    record.assignment_date.strftime("%Y-%m-%d") if record.assignment_date else "غير محدد",
                    record.completion_date.strftime("%Y-%m-%d") if record.completion_date else "مستمر",
                    record.project_manager or "غير محدد",
                    "نشط" if record.is_active else "منتهي",
                    record.notes or ""
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_projects.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = right_alignment
            
            # تعديل عرض الأعمدة
            column_widths = [25, 20, 15, 15, 20, 12, 25]
            for i, width in enumerate(column_widths, 1):
                ws_projects.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 5. ورقة سجلات التسليم والاستلام
        if handover_records and len(handover_records) > 0:
            ws_handover = wb.create_sheet("سجلات التسليم والاستلام")
            
            handover_headers = [
                "نوع العملية", "تاريخ العملية", "اسم الشخص", "الموظف", 
                "عداد المسافات", "حالة السيارة", "ملاحظات"
            ]
            
            # كتابة العناوين
            for col_idx, header in enumerate(handover_headers, 1):
                cell = ws_handover.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # كتابة البيانات
            for row_idx, record in enumerate(handover_records, 2):
                # الحصول على اسم الموظف
                employee_name = ""
                if record.employee_id:
                    from models import Employee
                    employee = Employee.query.get(record.employee_id)
                    if employee:
                        employee_name = employee.name
                
                row_data = [
                    "تسليم" if record.handover_type == "delivery" else "استلام",
                    record.handover_date.strftime("%Y-%m-%d") if record.handover_date else "غير محدد",
                    record.person_name or "غير محدد",
                    employee_name or "غير محدد",
                    str(record.mileage) if record.mileage else "غير محدد",
                    record.vehicle_condition or "غير محدد",
                    record.notes or ""
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_handover.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = right_alignment
            
            # تعديل عرض الأعمدة
            column_widths = [15, 15, 20, 20, 15, 20, 25]
            for i, width in enumerate(column_widths, 1):
                ws_handover.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 6. ورقة سجلات الفحص الدوري
        if inspection_records and len(inspection_records) > 0:
            ws_inspection = wb.create_sheet("سجلات الفحص الدوري")
            
            inspection_headers = [
                "تاريخ الفحص", "نوع الفحص", "نتيجة الفحص", "تاريخ الانتهاء", 
                "رقم الشهادة", "مركز الفحص", "التكلفة (ريال)", "ملاحظات"
            ]
            
            # كتابة العناوين
            for col_idx, header in enumerate(inspection_headers, 1):
                cell = ws_inspection.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # كتابة البيانات
            for row_idx, record in enumerate(inspection_records, 2):
                row_data = [
                    record.inspection_date.strftime("%Y-%m-%d") if record.inspection_date else "غير محدد",
                    record.inspection_type or "غير محدد",
                    "نجح" if record.result == "passed" else "فشل" if record.result == "failed" else "غير محدد",
                    record.expiry_date.strftime("%Y-%m-%d") if record.expiry_date else "غير محدد",
                    record.certificate_number or "غير محدد",
                    record.inspection_center or "غير محدد",
                    record.cost or 0,
                    record.notes or ""
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_inspection.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = right_alignment if col_idx != 7 else center_alignment
            
            # تعديل عرض الأعمدة
            column_widths = [15, 15, 15, 15, 20, 20, 12, 25]
            for i, width in enumerate(column_widths, 1):
                ws_inspection.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 7. ورقة فحوصات السلامة
        if safety_check_records and len(safety_check_records) > 0:
            ws_safety = wb.create_sheet("فحوصات السلامة")
            
            safety_headers = [
                "تاريخ الفحص", "نوع الفحص", "حالة الإطارات", "حالة الفرامل", 
                "حالة الأضواء", "مستوى الزيت", "الفاحص", "النتيجة العامة", "ملاحظات"
            ]
            
            # كتابة العناوين
            for col_idx, header in enumerate(safety_headers, 1):
                cell = ws_safety.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # كتابة البيانات
            for row_idx, record in enumerate(safety_check_records, 2):
                row_data = [
                    record.check_date.strftime("%Y-%m-%d") if record.check_date else "غير محدد",
                    record.check_type or "غير محدد",
                    record.tire_condition or "غير محدد",
                    record.brake_condition or "غير محدد",
                    record.lights_condition or "غير محدد",
                    record.oil_level or "غير محدد",
                    record.checked_by or "غير محدد",
                    record.overall_result or "غير محدد",
                    record.notes or ""
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_safety.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = right_alignment
            
            # تعديل عرض الأعمدة
            column_widths = [15, 15, 15, 15, 15, 15, 20, 15, 25]
            for i, width in enumerate(column_widths, 1):
                ws_safety.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 8. ورقة سجلات الحوادث
        if accident_records and len(accident_records) > 0:
            ws_accident = wb.create_sheet("سجلات الحوادث")
            
            accident_headers = [
                "تاريخ الحادث", "مكان الحادث", "وصف الحادث", "السائق", 
                "مستوى الضرر", "التكلفة المقدرة (ريال)", "حالة التأمين", "ملاحظات"
            ]
            
            # كتابة العناوين
            for col_idx, header in enumerate(accident_headers, 1):
                cell = ws_accident.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # كتابة البيانات
            for row_idx, record in enumerate(accident_records, 2):
                row_data = [
                    record.accident_date.strftime("%Y-%m-%d") if record.accident_date else "غير محدد",
                    record.location or "غير محدد",
                    record.description or "غير محدد",
                    record.driver_name or "غير محدد",
                    record.damage_level or "غير محدد",
                    record.estimated_cost or 0,
                    record.insurance_status or "غير محدد",
                    record.notes or ""
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_accident.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = right_alignment if col_idx != 6 else center_alignment
            
            # تعديل عرض الأعمدة
            column_widths = [15, 20, 25, 20, 15, 15, 15, 25]
            for i, width in enumerate(column_widths, 1):
                ws_accident.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # حفظ الملف
        wb.save(buffer)
        buffer.seek(0)
        
        print(f"تم إنشاء ملف Excel شامل بحجم: {len(buffer.getvalue())} بايت")
        return buffer
        
    except Exception as e:
        print(f"خطأ في تصدير البيانات الشاملة: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # إنشاء ملف بسيط في حالة الخطأ
        simple_data = {
            'البيان': ['رقم اللوحة', 'الماركة', 'الموديل', 'السائق الحالي', 'الحالة'],
            'القيمة': [
                vehicle.plate_number or '',
                vehicle.make or '',
                vehicle.model or '',
                vehicle.driver_name or "غير محدد",
                vehicle.status or ''
            ]
        }
        
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df_simple = pd.DataFrame(simple_data)
            df_simple.to_excel(writer, sheet_name='معلومات السيارة', index=False)
        
        buffer.seek(0)
        return buffer