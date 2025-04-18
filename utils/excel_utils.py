from openpyxl.utils import get_column_letter

def set_column_widths_safely(sheet):
    """تعيين أعراض الأعمدة بطريقة آمنة بدون استخدام الخلايا المدمجة"""
    try:
        # تعيين عرض افتراضي لكل الأعمدة
        for i in range(1, sheet.max_column + 1):
            column = get_column_letter(i)
            sheet.column_dimensions[column].width = 15
    except Exception as e:
        # تجاهل أي أخطاء قد تحدث أثناء ضبط الأعمدة
        print(f"خطأ عند ضبط عرض الأعمدة: {str(e)}")
    return sheet