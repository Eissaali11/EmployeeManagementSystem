"""
وحدة إنشاء تقارير Excel الشاملة للمركبات
توفر هذه الوحدة وظائف لإنشاء تقارير Excel شاملة للمركبات
"""

import io
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_complete_vehicle_excel_report(vehicle, rental=None, workshop_records=None, documents=None):
    """
    إنشاء تقرير شامل للسيارة بصيغة Excel
    
    Args:
        vehicle: كائن المركبة
        rental: معلومات الإيجار (اختياري)
        workshop_records: سجلات الورشة (اختياري)
        documents: مستندات المركبة (اختياري)
        
    Returns:
        bytes: محتوى ملف Excel
    """
    try:
        # إنشاء كائن جديد للمصنف
        workbook = Workbook()
        
        # تعريف الألوان والأنماط
        header_fill = PatternFill(start_color="1DA18E", end_color="1DA18E", fill_type="solid")
        subheader_fill = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")
        header_font = Font(name='Arial', bold=True, color="FFFFFF", size=12)
        subheader_font = Font(name='Arial', bold=True, size=11)
        normal_font = Font(name='Arial', size=10)
        
        # تعريف المحاذاة
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        # تعريف الحدود
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # الورقة الأولى - معلومات السيارة
        vehicle_sheet = workbook.active
        vehicle_sheet.title = "معلومات السيارة"
        
        # إضافة العنوان
        vehicle_sheet.merge_cells('A1:G1')
        title_cell = vehicle_sheet['A1']
        title_cell.value = f"تقرير شامل للسيارة: {vehicle.plate_number}"
        title_cell.font = Font(name='Arial', bold=True, size=14)
        title_cell.alignment = center_alignment
        title_cell.fill = PatternFill(start_color="1DA18E", end_color="1DA18E", fill_type="solid")
        title_cell.font = Font(name='Arial', bold=True, color="FFFFFF", size=14)
        
        # إضافة التاريخ
        vehicle_sheet.merge_cells('A2:G2')
        date_cell = vehicle_sheet['A2']
        date_cell.value = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d')}"
        date_cell.font = normal_font
        date_cell.alignment = center_alignment
        
        # ترويسة معلومات السيارة
        vehicle_sheet.merge_cells('A4:G4')
        vehicle_header = vehicle_sheet['A4']
        vehicle_header.value = "معلومات السيارة الأساسية"
        vehicle_header.font = subheader_font
        vehicle_header.alignment = right_alignment
        vehicle_header.fill = subheader_fill
        
        # تنسيق عرض الأعمدة
        for col in range(1, 8):
            vehicle_sheet.column_dimensions[get_column_letter(col)].width = 15
        
        # بيانات السيارة
        data_rows = [
            ["رقم اللوحة", vehicle.plate_number],
            ["النوع", f"{vehicle.make} {vehicle.model}"],
            ["سنة الصنع", str(vehicle.year)],
            ["اللون", vehicle.color],
            ["الحالة", _get_status_name(vehicle.status)],
            ["تاريخ الإضافة", vehicle.created_at.strftime('%Y-%m-%d')]
        ]
        
        # إضافة البيانات
        for row_idx, row_data in enumerate(data_rows, 5):
            vehicle_sheet.cell(row=row_idx, column=2).value = row_data[0]
            vehicle_sheet.cell(row=row_idx, column=2).font = normal_font
            vehicle_sheet.cell(row=row_idx, column=2).alignment = right_alignment
            vehicle_sheet.cell(row=row_idx, column=2).border = thin_border
            
            vehicle_sheet.cell(row=row_idx, column=3).value = row_data[1]
            vehicle_sheet.cell(row=row_idx, column=3).font = normal_font
            vehicle_sheet.cell(row=row_idx, column=3).alignment = right_alignment
            vehicle_sheet.cell(row=row_idx, column=3).border = thin_border
        
        # معلومات الإيجار
        rental_row = 12
        vehicle_sheet.merge_cells(f'A{rental_row}:G{rental_row}')
        rental_header = vehicle_sheet[f'A{rental_row}']
        rental_header.value = "معلومات الإيجار"
        rental_header.font = subheader_font
        rental_header.alignment = right_alignment
        rental_header.fill = subheader_fill
        
        rental_row += 1
        
        if rental:
            rental_data = [
                ["المؤجر", rental.lessor_name or "غير محدد"],
                ["تاريخ البداية", rental.start_date.strftime('%Y-%m-%d')],
                ["تاريخ النهاية", rental.end_date.strftime('%Y-%m-%d') if rental.end_date else "مستمر"],
                ["التكلفة الشهرية", f"{rental.monthly_cost:,.2f} ريال"],
                ["رقم العقد", rental.contract_number or "غير محدد"]
            ]
            
            for row_data in rental_data:
                vehicle_sheet.cell(row=rental_row, column=2).value = row_data[0]
                vehicle_sheet.cell(row=rental_row, column=2).font = normal_font
                vehicle_sheet.cell(row=rental_row, column=2).alignment = right_alignment
                vehicle_sheet.cell(row=rental_row, column=2).border = thin_border
                
                vehicle_sheet.cell(row=rental_row, column=3).value = row_data[1]
                vehicle_sheet.cell(row=rental_row, column=3).font = normal_font
                vehicle_sheet.cell(row=rental_row, column=3).alignment = right_alignment
                vehicle_sheet.cell(row=rental_row, column=3).border = thin_border
                
                rental_row += 1
        else:
            vehicle_sheet.merge_cells(f'A{rental_row}:G{rental_row}')
            no_rental_cell = vehicle_sheet[f'A{rental_row}']
            no_rental_cell.value = "لا يوجد إيجار نشط حاليًا"
            no_rental_cell.font = normal_font
            no_rental_cell.alignment = center_alignment
            
            rental_row += 2
        
        # ملخص سجلات الورشة
        workshop_row = rental_row + 2
        vehicle_sheet.merge_cells(f'A{workshop_row}:G{workshop_row}')
        workshop_header = vehicle_sheet[f'A{workshop_row}']
        workshop_header.value = "ملخص سجلات الورشة"
        workshop_header.font = subheader_font
        workshop_header.alignment = right_alignment
        workshop_header.fill = subheader_fill
        
        workshop_row += 1
        
        if workshop_records and len(workshop_records) > 0:
            # حساب التكلفة الإجمالية
            total_cost = sum(record.cost for record in workshop_records)
            
            workshop_summary = [
                ["عدد مرات دخول الورشة", str(len(workshop_records))],
                ["إجمالي التكاليف", f"{total_cost:,.2f} ريال"]
            ]
            
            for row_data in workshop_summary:
                vehicle_sheet.cell(row=workshop_row, column=2).value = row_data[0]
                vehicle_sheet.cell(row=workshop_row, column=2).font = normal_font
                vehicle_sheet.cell(row=workshop_row, column=2).alignment = right_alignment
                vehicle_sheet.cell(row=workshop_row, column=2).border = thin_border
                
                vehicle_sheet.cell(row=workshop_row, column=3).value = row_data[1]
                vehicle_sheet.cell(row=workshop_row, column=3).font = normal_font
                vehicle_sheet.cell(row=workshop_row, column=3).alignment = right_alignment
                vehicle_sheet.cell(row=workshop_row, column=3).border = thin_border
                
                workshop_row += 1
            
            # إنشاء ورقة سجلات الورشة
            workshop_sheet = workbook.create_sheet(title="سجلات الورشة")
            
            # ترويسة الجدول
            workshop_headers = ["تاريخ الدخول", "تاريخ الخروج", "سبب الدخول", "حالة الإصلاح", "اسم الورشة", "التكلفة", "ملاحظات"]
            
            for col_idx, header in enumerate(workshop_headers, 1):
                cell = workshop_sheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
                workshop_sheet.column_dimensions[get_column_letter(col_idx)].width = 20
            
            # بيانات سجلات الورشة
            for row_idx, record in enumerate(workshop_records, 2):
                workshop_sheet.cell(row=row_idx, column=1).value = record.entry_date.strftime('%Y-%m-%d')
                workshop_sheet.cell(row=row_idx, column=2).value = record.exit_date.strftime('%Y-%m-%d') if record.exit_date else "ما زالت في الورشة"
                workshop_sheet.cell(row=row_idx, column=3).value = _get_reason_name(record.reason)
                workshop_sheet.cell(row=row_idx, column=4).value = _get_repair_status_name(record.repair_status)
                workshop_sheet.cell(row=row_idx, column=5).value = record.workshop_name or "غير محدد"
                workshop_sheet.cell(row=row_idx, column=6).value = f"{record.cost:,.2f}"
                workshop_sheet.cell(row=row_idx, column=7).value = record.notes or ""
                
                # تنسيق الخلايا
                for col_idx in range(1, 8):
                    cell = workshop_sheet.cell(row=row_idx, column=col_idx)
                    cell.font = normal_font
                    cell.alignment = right_alignment
                    cell.border = thin_border
        else:
            vehicle_sheet.merge_cells(f'A{workshop_row}:G{workshop_row}')
            no_workshop_cell = vehicle_sheet[f'A{workshop_row}']
            no_workshop_cell.value = "لا توجد سجلات ورشة لهذه السيارة"
            no_workshop_cell.font = normal_font
            no_workshop_cell.alignment = center_alignment
        
        # إضافة الملاحظات إن وجدت
        if vehicle.notes:
            notes_row = workshop_row + 2
            vehicle_sheet.merge_cells(f'A{notes_row}:G{notes_row}')
            notes_header = vehicle_sheet[f'A{notes_row}']
            notes_header.value = "ملاحظات"
            notes_header.font = subheader_font
            notes_header.alignment = right_alignment
            notes_header.fill = subheader_fill
            
            notes_row += 1
            vehicle_sheet.merge_cells(f'A{notes_row}:G{notes_row + 3}')
            notes_cell = vehicle_sheet[f'A{notes_row}']
            notes_cell.value = vehicle.notes
            notes_cell.font = normal_font
            notes_cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
        
        # إرجاع ملف Excel كـ bytes
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
        
    except Exception as e:
        print(f"خطأ في إنشاء تقرير Excel الشامل للسيارة: {str(e)}")
        raise e


def _get_status_name(status):
    """تحويل حالة السيارة إلى النص العربي المناسب"""
    status_map = {
        'available': 'متاحة',
        'rented': 'مؤجرة',
        'in_project': 'في المشروع',
        'in_workshop': 'في الورشة',
        'accident': 'حادث'
    }
    return status_map.get(status, status)


def _get_reason_name(reason):
    """تحويل سبب دخول الورشة إلى النص العربي المناسب"""
    reason_map = {
        'maintenance': 'صيانة دورية',
        'breakdown': 'عطل',
        'accident': 'حادث'
    }
    return reason_map.get(reason, reason)


def _get_repair_status_name(status):
    """تحويل حالة الإصلاح إلى النص العربي المناسب"""
    status_map = {
        'in_progress': 'قيد التنفيذ',
        'completed': 'تم الإصلاح',
        'pending_approval': 'بانتظار الموافقة'
    }
    return status_map.get(status, status)