"""
نظام تصدير Excel شامل للسيارات مع تبويبات منفصلة
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
from datetime import datetime


def create_vehicle_excel_with_tabs(vehicle, 
                                 workshop_records=None,
                                 rental_records=None,
                                 project_records=None,
                                 handover_records=None,
                                 inspection_records=None,
                                 safety_check_records=None,
                                 accident_records=None):
    """
    إنشاء ملف Excel شامل للسيارة مع تبويبات منفصلة لكل قسم
    """
    try:
        print(f"Starting tabbed export for vehicle: {vehicle.plate_number if vehicle else 'None'}")
        print(f"Workshop records: {len(workshop_records) if workshop_records else 0}")
        print(f"Rental records: {len(rental_records) if rental_records else 0}")
        # إنشاء كتاب العمل
        wb = openpyxl.Workbook()
        
        # حذف الورقة الافتراضية
        wb.remove(wb.active)
        
        # إعداد الأنماط
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        data_font = Font(name='Arial', size=10)
        center_alignment = Alignment(horizontal='center', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # 1. تبويب معلومات السيارة الأساسية
        ws_basic = wb.create_sheet("معلومات السيارة")
        
        # عنوان الصفحة
        ws_basic.merge_cells('A1:B1')
        title_cell = ws_basic['A1']
        title_cell.value = f"معلومات السيارة - {vehicle.plate_number}"
        title_cell.font = Font(name='Arial', size=14, bold=True)
        title_cell.alignment = center_alignment
        title_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # البيانات الأساسية
        basic_data = [
            ["البيان", "القيمة"],
            ["رقم اللوحة", vehicle.plate_number or ""],
            ["الماركة", vehicle.make or ""],
            ["الموديل", vehicle.model or ""],
            ["سنة الصنع", str(vehicle.year) if vehicle.year else ""],
            ["اللون", vehicle.color or ""],
            ["السائق الحالي", vehicle.driver_name or "غير محدد"],
            ["الحالة", {
                'available': 'متاحة',
                'rented': 'مؤجرة', 
                'in_workshop': 'في الورشة',
                'in_project': 'في المشروع',
                'accident': 'حادث',
                'sold': 'مباعة'
            }.get(vehicle.status, vehicle.status or "")],
            ["تاريخ انتهاء التفويض", vehicle.authorization_expiry_date.strftime("%Y-%m-%d") if vehicle.authorization_expiry_date else "غير محدد"],
            ["تاريخ انتهاء الاستمارة", vehicle.registration_expiry_date.strftime("%Y-%m-%d") if vehicle.registration_expiry_date else "غير محدد"],
            ["تاريخ انتهاء الفحص الدوري", vehicle.inspection_expiry_date.strftime("%Y-%m-%d") if vehicle.inspection_expiry_date else "غير محدد"],
            ["الملاحظات", vehicle.notes or ""],
            ["تاريخ الإضافة", vehicle.created_at.strftime("%Y-%m-%d %H:%M") if vehicle.created_at else ""],
            ["آخر تحديث", vehicle.updated_at.strftime("%Y-%m-%d %H:%M") if vehicle.updated_at else ""]
        ]
        
        # كتابة البيانات الأساسية
        for row_idx, row_data in enumerate(basic_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_basic.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.font = header_font if row_idx == 2 else data_font
                cell.alignment = right_alignment
                if row_idx == 2:
                    cell.fill = header_fill
        
        # تعديل عرض الأعمدة
        ws_basic.column_dimensions['A'].width = 25
        ws_basic.column_dimensions['B'].width = 30
        
        # 2. تبويب سجلات الورشة
        if workshop_records:
            ws_workshop = wb.create_sheet("سجلات الورشة")
            
            # عنوان الصفحة
            ws_workshop.merge_cells('A1:H1')
            title_cell = ws_workshop['A1']
            title_cell.value = f"سجلات الورشة - {vehicle.plate_number}"
            title_cell.font = Font(name='Arial', size=14, bold=True)
            title_cell.alignment = center_alignment
            title_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            # العناوين
            headers = ["تاريخ الدخول", "تاريخ الخروج", "نوع الصيانة", "الوصف", "التكلفة", "الحالة", "الورشة", "الملاحظات"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws_workshop.cell(row=2, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_alignment
            
            # البيانات
            for row_idx, record in enumerate(workshop_records, 3):
                data_row = [
                    record.entry_date.strftime("%Y-%m-%d") if record.entry_date else "",
                    record.exit_date.strftime("%Y-%m-%d") if record.exit_date else "لم يخرج بعد",
                    record.reason or "",
                    record.description or "",
                    f"{record.cost:.2f}" if record.cost else "",
                    {'pending': 'قيد الانتظار', 'in_progress': 'قيد التنفيذ', 'completed': 'مكتمل', 'cancelled': 'ملغي'}.get(record.repair_status, record.repair_status or ""),
                    record.workshop_name or "",
                    record.notes or ""
                ]
                
                for col_idx, value in enumerate(data_row, 1):
                    cell = ws_workshop.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = right_alignment
                    cell.font = data_font
            
            # تعديل عرض الأعمدة
            for i in range(1, 9):
                column_letter = chr(64 + i)  # A, B, C, etc.
                ws_workshop.column_dimensions[column_letter].width = 15
        
        # 3. تبويب سجلات الإيجار
        if rental_records:
            ws_rental = wb.create_sheet("سجلات الإيجار")
            
            # عنوان الصفحة
            ws_rental.merge_cells('A1:G1')
            title_cell = ws_rental['A1']
            title_cell.value = f"سجلات الإيجار - {vehicle.plate_number}"
            title_cell.font = Font(name='Arial', size=14, bold=True)
            title_cell.alignment = center_alignment
            title_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            # العناوين
            headers = ["تاريخ البداية", "تاريخ النهاية", "المستأجر", "مبلغ الإيجار", "مبلغ التأمين", "الحالة", "الملاحظات"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws_rental.cell(row=2, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_alignment
            
            # البيانات
            for row_idx, record in enumerate(rental_records, 3):
                data_row = [
                    record.start_date.strftime("%Y-%m-%d") if record.start_date else "",
                    record.end_date.strftime("%Y-%m-%d") if record.end_date else "",
                    record.renter_name or "",
                    f"{record.rental_amount:.2f}" if record.rental_amount else "",
                    f"{record.security_deposit:.2f}" if record.security_deposit else "",
                    {'active': 'نشط', 'completed': 'مكتمل', 'cancelled': 'ملغي'}.get(record.status, record.status or ""),
                    record.notes or ""
                ]
                
                for col_idx, value in enumerate(data_row, 1):
                    cell = ws_rental.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = right_alignment
                    cell.font = data_font
            
            # تعديل عرض الأعمدة
            for i in range(1, 8):
                column_letter = chr(64 + i)
                ws_rental.column_dimensions[column_letter].width = 15
        
        # 4. تبويب سجلات المشاريع
        if project_records:
            ws_project = wb.create_sheet("سجلات المشاريع")
            
            # عنوان الصفحة
            ws_project.merge_cells('A1:G1')
            title_cell = ws_project['A1']
            title_cell.value = f"سجلات المشاريع - {vehicle.plate_number}"
            title_cell.font = Font(name='Arial', size=14, bold=True)
            title_cell.alignment = center_alignment
            title_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            # العناوين
            headers = ["اسم المشروع", "الموقع", "مدير المشروع", "تاريخ البداية", "تاريخ النهاية", "الحالة", "الملاحظات"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws_project.cell(row=2, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_alignment
            
            # البيانات
            for row_idx, record in enumerate(project_records, 3):
                data_row = [
                    record.project_name or "",
                    record.location or "",
                    record.manager_name or "",
                    record.start_date.strftime("%Y-%m-%d") if record.start_date else "",
                    record.end_date.strftime("%Y-%m-%d") if record.end_date else "مستمر",
                    "نشط" if record.is_active else "غير نشط",
                    record.notes or ""
                ]
                
                for col_idx, value in enumerate(data_row, 1):
                    cell = ws_project.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = right_alignment
                    cell.font = data_font
            
            # تعديل عرض الأعمدة
            for i in range(1, 8):
                column_letter = chr(64 + i)
                ws_project.column_dimensions[column_letter].width = 18
        
        # 5. تبويب سجلات التسليم والاستلام
        if handover_records:
            ws_handover = wb.create_sheet("التسليم والاستلام")
            
            # عنوان الصفحة
            ws_handover.merge_cells('A1:I1')
            title_cell = ws_handover['A1']
            title_cell.value = f"سجلات التسليم والاستلام - {vehicle.plate_number}"
            title_cell.font = Font(name='Arial', size=14, bold=True)
            title_cell.alignment = center_alignment
            title_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            # العناوين
            headers = ["النوع", "التاريخ", "اسم المستلم/المسلم", "المشرف", "حالة السيارة", "مستوى الوقود", "عداد الكيلومترات", "الملاحظات", "رابط النموذج"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws_handover.cell(row=2, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_alignment
            
            # البيانات
            for row_idx, record in enumerate(handover_records, 3):
                data_row = [
                    {'delivery': 'تسليم', 'return': 'استلام'}.get(record.handover_type, record.handover_type or ""),
                    record.handover_date.strftime("%Y-%m-%d") if record.handover_date else "",
                    record.person_name or "",
                    record.supervisor_name or "",
                    record.vehicle_condition or "",
                    record.fuel_level or "",
                    str(record.mileage) if record.mileage else "",
                    record.notes or "",
                    record.form_link or ""
                ]
                
                for col_idx, value in enumerate(data_row, 1):
                    cell = ws_handover.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = right_alignment
                    cell.font = data_font
            
            # تعديل عرض الأعمدة
            for i in range(1, 10):
                column_letter = chr(64 + i)
                ws_handover.column_dimensions[column_letter].width = 15
        
        # حفظ الملف في الذاكرة
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        print(f"خطأ في إنشاء ملف التصدير الشامل: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # إنشاء ملف بسيط في حالة الخطأ
        wb_simple = openpyxl.Workbook()
        ws_simple = wb_simple.active
        ws_simple.title = "معلومات السيارة"
        
        # إضافة البيانات الأساسية
        headers = ['البيان', 'القيمة']
        data = [
            ['رقم اللوحة', vehicle.plate_number or ''],
            ['الماركة', vehicle.make or ''],
            ['الموديل', vehicle.model or ''],
            ['السائق الحالي', vehicle.driver_name or "غير محدد"],
            ['الحالة', vehicle.status or '']
        ]
        
        # كتابة العناوين
        for col_idx, header in enumerate(headers, 1):
            ws_simple.cell(row=1, column=col_idx, value=header)
        
        # كتابة البيانات
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws_simple.cell(row=row_idx, column=col_idx, value=value)
        
        buffer = BytesIO()
        wb_simple.save(buffer)
        buffer.seek(0)
        return buffer