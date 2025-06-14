"""
وحدة للتصدير والمشاركة في نظام إدارة المركبات
تتضمن وظائف لتصدير بيانات المركبات وسجلات الورشة إلى PDF وExcel
"""

import os
import io
import tempfile
from datetime import datetime
from flask import url_for
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from arabic_reshaper import reshape
from bidi.algorithm import get_display
import pandas as pd

# تسجيل الخطوط العربية
def register_fonts():
    """تسجيل الخطوط العربية"""
    font_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'static', 'fonts')
    
    # التحقق من وجود الخطوط وتسجيلها
    amiri_path = os.path.join(font_path, 'Amiri-Regular.ttf')
    amiri_bold_path = os.path.join(font_path, 'Amiri-Bold.ttf')
    
    if not os.path.exists(font_path):
        os.makedirs(font_path)
    
    if os.path.exists(amiri_path) and os.path.exists(amiri_bold_path):
        pdfmetrics.registerFont(TTFont('Amiri', amiri_path))
        pdfmetrics.registerFont(TTFont('Amiri-Bold', amiri_bold_path))
    else:
        # استخدام خط افتراضي إذا لم تكن الخطوط العربية متوفرة
        pdfmetrics.registerFont(TTFont('Amiri', 'Helvetica'))
        pdfmetrics.registerFont(TTFont('Amiri-Bold', 'Helvetica-Bold'))


def arabic_text(text):
    """معالجة النص العربي للعرض الصحيح في ملفات PDF"""
    if text is None:
        return ""
    return get_display(reshape(str(text)))


def export_vehicle_pdf(vehicle, workshop_records=None, rental_records=None):
    """
    تصدير بيانات السيارة إلى ملف PDF
    
    Args:
        vehicle: كائن السيارة
        workshop_records: سجلات الورشة (اختياري)
        rental_records: سجلات الإيجار (اختياري)
    
    Returns:
        BytesIO: كائن بايت يحتوي على ملف PDF
    """
    register_fonts()
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    styles = getSampleStyleSheet()
    
    # إنشاء نمط للنص العربي
    styles.add(ParagraphStyle(name='Arabic', fontName='Amiri', fontSize=12, alignment=1))  # alignment=1 للتوسيط
    styles.add(ParagraphStyle(name='ArabicTitle', fontName='Amiri-Bold', fontSize=16, alignment=1))
    styles.add(ParagraphStyle(name='ArabicHeading', fontName='Amiri-Bold', fontSize=14, alignment=1))
    
    # تحضير المحتوى
    content = []
    
    # إضافة شعار الشركة إذا كان متوفراً
    logo_path = os.path.join(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "static"), 'img/logo.png')
    if os.path.exists(logo_path):
        img = Image(logo_path, width=120, height=60)
        content.append(img)
        content.append(Spacer(1, 10))
    
    # عنوان التقرير
    title = Paragraph(arabic_text(f"تقرير بيانات السيارة: {vehicle.plate_number}"), styles['ArabicTitle'])
    content.append(title)
    content.append(Spacer(1, 20))
    
    # معلومات السيارة الأساسية
    content.append(Paragraph(arabic_text("معلومات السيارة الأساسية"), styles['ArabicHeading']))
    content.append(Spacer(1, 10))
    
    # جدول المعلومات الأساسية
    basic_data = [
        [arabic_text("رقم اللوحة"), arabic_text(vehicle.plate_number)],
        [arabic_text("النوع"), arabic_text(f"{vehicle.make} {vehicle.model}")],
        [arabic_text("سنة الصنع"), arabic_text(str(vehicle.year))],
        [arabic_text("اللون"), arabic_text(vehicle.color)],
        [arabic_text("السائق الحالي"), arabic_text(vehicle.driver_name or "غير محدد")],
        [arabic_text("القسم/المالك"), arabic_text(vehicle.department or "غير محدد")],
        [arabic_text("الحالة"), arabic_text({
            'available': 'متاحة',
            'rented': 'مؤجرة',
            'in_workshop': 'في الورشة',
            'in_project': 'في المشروع',
            'accident': 'حادث',
            'sold': 'مباعة'
        }.get(vehicle.status, vehicle.status))],
        [arabic_text("تاريخ انتهاء التأمين"), arabic_text(vehicle.insurance_expiry.strftime("%Y-%m-%d") if vehicle.insurance_expiry else "غير محدد")],
        [arabic_text("تاريخ انتهاء الفحص"), arabic_text(vehicle.inspection_expiry.strftime("%Y-%m-%d") if vehicle.inspection_expiry else "غير محدد")],
    ]
    
    t = Table(basic_data, colWidths=[doc.width/3, 2*doc.width/3])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Amiri'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    content.append(t)
    content.append(Spacer(1, 20))
    
    # سجلات الورشة إذا كانت متوفرة
    if workshop_records and len(workshop_records) > 0:
        content.append(Paragraph(arabic_text("سجلات الورشة"), styles['ArabicHeading']))
        content.append(Spacer(1, 10))
        
        # تحضير بيانات سجلات الورشة
        workshop_data = [
            [
                arabic_text("سبب الدخول"),
                arabic_text("تاريخ الدخول"),
                arabic_text("تاريخ الخروج"),
                arabic_text("حالة الإصلاح"),
                arabic_text("التكلفة (ريال)"),
                arabic_text("اسم الورشة")
            ]
        ]
        
        for record in workshop_records:
            reason_map = {'maintenance': 'صيانة دورية', 'breakdown': 'عطل', 'accident': 'حادث'}
            status_map = {'in_progress': 'قيد التنفيذ', 'completed': 'تم الإصلاح', 'pending_approval': 'بانتظار الموافقة'}
            
            workshop_data.append([
                arabic_text(reason_map.get(record.reason, record.reason)),
                arabic_text(record.entry_date.strftime("%Y-%m-%d")),
                arabic_text(record.exit_date.strftime("%Y-%m-%d") if record.exit_date else "ما زالت في الورشة"),
                arabic_text(status_map.get(record.repair_status, record.repair_status)),
                arabic_text(f"{record.cost:,.2f}"),
                arabic_text(record.workshop_name or "غير محدد")
            ])
        
        t = Table(workshop_data, colWidths=[doc.width/6] * 6)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (4, 1), (4, -1), 'LEFT'),  # محاذاة أرقام التكلفة إلى اليسار
            ('FONTNAME', (0, 0), (-1, -1), 'Amiri'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        content.append(t)
        content.append(Spacer(1, 10))
        
        # إجمالي تكاليف الصيانة
        total_cost = sum(record.cost for record in workshop_records)
        cost_text = Paragraph(arabic_text(f"إجمالي تكاليف الصيانة: {total_cost:,.2f} ريال"), styles['Arabic'])
        content.append(cost_text)
        content.append(Spacer(1, 20))
    
    # سجلات الإيجار إذا كانت متوفرة
    if rental_records and len(rental_records) > 0:
        content.append(Paragraph(arabic_text("سجلات الإيجار"), styles['ArabicHeading']))
        content.append(Spacer(1, 10))
        
        # تحضير بيانات سجلات الإيجار
        rental_data = [
            [
                arabic_text("المستأجر"),
                arabic_text("تاريخ البداية"),
                arabic_text("تاريخ النهاية"),
                arabic_text("التكلفة (ريال)"),
                arabic_text("الحالة")
            ]
        ]
        
        for record in rental_records:
            rental_data.append([
                arabic_text(record.renter_name),
                arabic_text(record.start_date.strftime("%Y-%m-%d")),
                arabic_text(record.end_date.strftime("%Y-%m-%d") if record.end_date else "مستمر"),
                arabic_text(f"{record.cost:,.2f}"),
                arabic_text("نشط" if record.is_active else "منتهي")
            ])
        
        t = Table(rental_data, colWidths=[doc.width/5] * 5)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (3, 1), (3, -1), 'LEFT'),  # محاذاة أرقام التكلفة إلى اليسار
            ('FONTNAME', (0, 0), (-1, -1), 'Amiri'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        content.append(t)
        content.append(Spacer(1, 20))
    
    # بيانات التوقيع والطباعة
    footer_text = Paragraph(
        arabic_text(f"تم إنشاء هذا التقرير بواسطة نُظم - نظام إدارة متكامل في {datetime.now().strftime('%Y-%m-%d %H:%M')}"),
        styles['Arabic']
    )
    content.append(footer_text)
    
    # بناء الوثيقة
    doc.build(content)
    buffer.seek(0)
    return buffer


def export_workshop_records_pdf(vehicle, workshop_records):
    """
    تصدير سجلات الورشة لسيارة معينة إلى ملف PDF
    
    Args:
        vehicle: كائن السيارة
        workshop_records: سجلات الورشة
    
    Returns:
        BytesIO: كائن بايت يحتوي على ملف PDF
    """
    register_fonts()
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    styles = getSampleStyleSheet()
    
    # إنشاء نمط للنص العربي
    styles.add(ParagraphStyle(name='Arabic', fontName='Amiri', fontSize=12, alignment=1))
    styles.add(ParagraphStyle(name='ArabicTitle', fontName='Amiri-Bold', fontSize=16, alignment=1))
    styles.add(ParagraphStyle(name='ArabicHeading', fontName='Amiri-Bold', fontSize=14, alignment=1))
    
    # تحضير المحتوى
    content = []
    
    # إضافة شعار الشركة إذا كان متوفراً
    logo_path = os.path.join(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "static"), 'img/logo.png')
    if os.path.exists(logo_path):
        img = Image(logo_path, width=120, height=60)
        content.append(img)
        content.append(Spacer(1, 10))
    
    # عنوان التقرير
    title = Paragraph(arabic_text(f"تقرير سجلات الورشة للسيارة: {vehicle.plate_number}"), styles['ArabicTitle'])
    content.append(title)
    content.append(Spacer(1, 10))
    
    # معلومات السيارة الأساسية
    vehicle_info = Paragraph(
        arabic_text(f"السيارة: {vehicle.make} {vehicle.model} {vehicle.year} - {vehicle.color}"),
        styles['Arabic']
    )
    content.append(vehicle_info)
    content.append(Spacer(1, 20))
    
    # سجلات الورشة
    content.append(Paragraph(arabic_text("سجلات الورشة"), styles['ArabicHeading']))
    content.append(Spacer(1, 10))
    
    if workshop_records and len(workshop_records) > 0:
        # تحضير بيانات سجلات الورشة
        workshop_data = [
            [
                arabic_text("سبب الدخول"),
                arabic_text("تاريخ الدخول"),
                arabic_text("تاريخ الخروج"),
                arabic_text("حالة الإصلاح"),
                arabic_text("التكلفة (ريال)"),
                arabic_text("اسم الورشة"),
                arabic_text("الفني المسؤول")
            ]
        ]
        
        for record in workshop_records:
            reason_map = {'maintenance': 'صيانة دورية', 'breakdown': 'عطل', 'accident': 'حادث'}
            status_map = {'in_progress': 'قيد التنفيذ', 'completed': 'تم الإصلاح', 'pending_approval': 'بانتظار الموافقة'}
            
            workshop_data.append([
                arabic_text(reason_map.get(record.reason, record.reason)),
                arabic_text(record.entry_date.strftime("%Y-%m-%d")),
                arabic_text(record.exit_date.strftime("%Y-%m-%d") if record.exit_date else "ما زالت في الورشة"),
                arabic_text(status_map.get(record.repair_status, record.repair_status)),
                arabic_text(f"{record.cost:,.2f}"),
                arabic_text(record.workshop_name or "غير محدد"),
                arabic_text(record.technician_name or "غير محدد")
            ])
        
        t = Table(workshop_data, colWidths=[doc.width/7] * 7)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (4, 1), (4, -1), 'LEFT'),  # محاذاة أرقام التكلفة إلى اليسار
            ('FONTNAME', (0, 0), (-1, -1), 'Amiri'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        content.append(t)
        content.append(Spacer(1, 20))
        
        # إجمالي تكاليف الصيانة
        total_cost = sum(record.cost for record in workshop_records)
        cost_text = Paragraph(arabic_text(f"إجمالي تكاليف الصيانة: {total_cost:,.2f} ريال"), styles['Arabic'])
        content.append(cost_text)
        
        # إجمالي عدد أيام الورشة
        days_in_workshop = sum(
            (record.exit_date - record.entry_date).days + 1 if record.exit_date else 
            (datetime.now().date() - record.entry_date).days + 1
            for record in workshop_records
        )
        days_text = Paragraph(arabic_text(f"إجمالي عدد أيام الورشة: {days_in_workshop} يوم"), styles['Arabic'])
        content.append(days_text)
    else:
        content.append(Paragraph(arabic_text("لا توجد سجلات ورشة لهذه السيارة"), styles['Arabic']))
    
    content.append(Spacer(1, 20))
    
    # بيانات التوقيع والطباعة
    footer_text = Paragraph(
        arabic_text(f"تم إنشاء هذا التقرير بواسطة نُظم - نظام إدارة متكامل في {datetime.now().strftime('%Y-%m-%d %H:%M')}"),
        styles['Arabic']
    )
    content.append(footer_text)
    
    # بناء الوثيقة
    doc.build(content)
    buffer.seek(0)
    return buffer


def export_vehicle_excel(vehicle, workshop_records=None, rental_records=None):
    """
    تصدير بيانات السيارة إلى ملف Excel
    
    Args:
        vehicle: كائن السيارة
        workshop_records: سجلات الورشة (اختياري)
        rental_records: سجلات الإيجار (اختياري)
    
    Returns:
        BytesIO: كائن بايت يحتوي على ملف Excel
    """
    buffer = io.BytesIO()
    
    # إنشاء مصنف Excel مع عدة أوراق عمل
    with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
        workbook = writer.book
        
        # نمط العناوين
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'center',
            'align': 'center',
            'fg_color': '#D7E4BC',
            'border': 1
        })
        
        # نمط الخلايا
        cell_format = workbook.add_format({
            'align': 'right',
            'valign': 'vcenter',
            'border': 1
        })
        
        # ورقة المعلومات الأساسية
        basic_data = {
            'البيان': [
                'رقم اللوحة', 'النوع', 'سنة الصنع', 'اللون', 'اسم السائق',
                'الحالة', 'تاريخ انتهاء التفويض', 'تاريخ انتهاء الاستمارة', 'تاريخ انتهاء الفحص الدوري'
            ],
            'القيمة': [
                vehicle.plate_number,
                f"{vehicle.make} {vehicle.model}",
                str(vehicle.year),
                vehicle.color,
                vehicle.driver_name or "غير محدد",
                {
                    'available': 'متاحة',
                    'rented': 'مؤجرة',
                    'in_workshop': 'في الورشة',
                    'in_project': 'في المشروع',
                    'accident': 'حادث',
                    'sold': 'مباعة'
                }.get(vehicle.status, vehicle.status),
                vehicle.authorization_expiry_date.strftime("%Y-%m-%d") if vehicle.authorization_expiry_date else "غير محدد",
                vehicle.registration_expiry_date.strftime("%Y-%m-%d") if vehicle.registration_expiry_date else "غير محدد",
                vehicle.inspection_expiry_date.strftime("%Y-%m-%d") if vehicle.inspection_expiry_date else "غير محدد"
            ]
        }
        
        # إنشاء DataFrame وكتابته إلى ورقة العمل
        df_basic = pd.DataFrame(basic_data)
        df_basic.to_excel(writer, sheet_name='معلومات السيارة', index=False)
        
        # تعديل عرض الأعمدة
        worksheet = writer.sheets['معلومات السيارة']
        worksheet.set_column('A:A', 25)
        worksheet.set_column('B:B', 30)
        
        # تنسيق الخلايا
        for col_num, col in enumerate(df_basic.columns):
            worksheet.write(0, col_num, col, header_format)
        
        for row_num in range(len(df_basic)):
            for col_num in range(len(df_basic.columns)):
                worksheet.write(row_num + 1, col_num, df_basic.iloc[row_num, col_num], cell_format)
        
        # إذا كانت سجلات الورشة متوفرة
        if workshop_records and len(workshop_records) > 0:
            # تحويل سجلات الورشة إلى DataFrame
            workshop_data = []
            
            for record in workshop_records:
                reason_map = {'maintenance': 'صيانة دورية', 'breakdown': 'عطل', 'accident': 'حادث'}
                status_map = {'in_progress': 'قيد التنفيذ', 'completed': 'تم الإصلاح', 'pending_approval': 'بانتظار الموافقة'}
                
                workshop_data.append({
                    'سبب الدخول': reason_map.get(record.reason, record.reason),
                    'تاريخ الدخول': record.entry_date.strftime("%Y-%m-%d"),
                    'تاريخ الخروج': record.exit_date.strftime("%Y-%m-%d") if record.exit_date else "ما زالت في الورشة",
                    'حالة الإصلاح': status_map.get(record.repair_status, record.repair_status),
                    'التكلفة (ريال)': record.cost,
                    'اسم الورشة': record.workshop_name or "غير محدد",
                    'الفني المسؤول': record.technician_name or "غير محدد",
                    'الوصف': record.description or "",
                    'ملاحظات': record.notes or ""
                })
            
            df_workshop = pd.DataFrame(workshop_data)
            df_workshop.to_excel(writer, sheet_name='سجلات الورشة', index=False)
            
            # تعديل عرض الأعمدة
            worksheet = writer.sheets['سجلات الورشة']
            worksheet.set_column('A:A', 15)
            worksheet.set_column('B:C', 15)
            worksheet.set_column('D:D', 15)
            worksheet.set_column('E:E', 15)
            worksheet.set_column('F:G', 20)
            worksheet.set_column('H:I', 30)
            
            # تنسيق الخلايا
            for col_num, col in enumerate(df_workshop.columns):
                worksheet.write(0, col_num, col, header_format)
            
            for row_num in range(len(df_workshop)):
                for col_num in range(len(df_workshop.columns)):
                    worksheet.write(row_num + 1, col_num, df_workshop.iloc[row_num, col_num], cell_format)
            
            # إضافة صف إجمالي
            total_row = len(df_workshop) + 2
            worksheet.write(total_row, 0, 'الإجمالي', header_format)
            worksheet.write_formula(total_row, 4, f'=SUM(E2:E{len(df_workshop) + 1})', header_format)
            
            # دمج الخلايا للإجمالي
            worksheet.merge_range(total_row, 0, total_row, 3, 'الإجمالي', header_format)
        
        # إذا كانت سجلات الإيجار متوفرة
        if rental_records and len(rental_records) > 0:
            # تحويل سجلات الإيجار إلى DataFrame
            rental_data = []
            
            for record in rental_records:
                rental_data.append({
                    'المستأجر': record.renter_name,
                    'تاريخ البداية': record.start_date.strftime("%Y-%m-%d"),
                    'تاريخ النهاية': record.end_date.strftime("%Y-%m-%d") if record.end_date else "مستمر",
                    'التكلفة (ريال)': record.cost,
                    'الحالة': "نشط" if record.is_active else "منتهي",
                    'جهة الاتصال': record.contact_number or "",
                    'ملاحظات': record.notes or ""
                })
            
            df_rental = pd.DataFrame(rental_data)
            df_rental.to_excel(writer, sheet_name='سجلات الإيجار', index=False)
            
            # تعديل عرض الأعمدة
            worksheet = writer.sheets['سجلات الإيجار']
            worksheet.set_column('A:A', 20)
            worksheet.set_column('B:C', 15)
            worksheet.set_column('D:D', 15)
            worksheet.set_column('E:E', 10)
            worksheet.set_column('F:F', 15)
            worksheet.set_column('G:G', 30)
            
            # تنسيق الخلايا
            for col_num, col in enumerate(df_rental.columns):
                worksheet.write(0, col_num, col, header_format)
            
            for row_num in range(len(df_rental)):
                for col_num in range(len(df_rental.columns)):
                    worksheet.write(row_num + 1, col_num, df_rental.iloc[row_num, col_num], cell_format)
    
    buffer.seek(0)
    return buffer


def export_workshop_records_excel(vehicle, workshop_records):
    """
    تصدير سجلات الورشة لسيارة معينة إلى ملف Excel
    
    Args:
        vehicle: كائن السيارة
        workshop_records: سجلات الورشة
    
    Returns:
        BytesIO: كائن بايت يحتوي على ملف Excel
    """
    buffer = io.BytesIO()
    
    # إنشاء مصنف Excel
    with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
        workbook = writer.book
        
        # نمط العناوين
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'center',
            'align': 'center',
            'fg_color': '#D7E4BC',
            'border': 1
        })
        
        # نمط الخلايا
        cell_format = workbook.add_format({
            'align': 'right',
            'valign': 'vcenter',
            'border': 1
        })
        
        # معلومات السيارة
        vehicle_info = {
            'البيان': ['رقم اللوحة', 'النوع', 'سنة الصنع', 'اللون'],
            'القيمة': [
                vehicle.plate_number,
                f"{vehicle.make} {vehicle.model}",
                str(vehicle.year),
                vehicle.color
            ]
        }
        
        df_vehicle = pd.DataFrame(vehicle_info)
        df_vehicle.to_excel(writer, sheet_name='معلومات السيارة', index=False)
        
        # تعديل عرض الأعمدة
        worksheet = writer.sheets['معلومات السيارة']
        worksheet.set_column('A:A', 15)
        worksheet.set_column('B:B', 25)
        
        # تنسيق الخلايا
        for col_num, col in enumerate(df_vehicle.columns):
            worksheet.write(0, col_num, col, header_format)
        
        for row_num in range(len(df_vehicle)):
            for col_num in range(len(df_vehicle.columns)):
                worksheet.write(row_num + 1, col_num, df_vehicle.iloc[row_num, col_num], cell_format)
        
        # سجلات الورشة
        if workshop_records and len(workshop_records) > 0:
            # تحويل سجلات الورشة إلى DataFrame
            workshop_data = []
            
            for record in workshop_records:
                reason_map = {'maintenance': 'صيانة دورية', 'breakdown': 'عطل', 'accident': 'حادث'}
                status_map = {'in_progress': 'قيد التنفيذ', 'completed': 'تم الإصلاح', 'pending_approval': 'بانتظار الموافقة'}
                
                workshop_data.append({
                    'سبب الدخول': reason_map.get(record.reason, record.reason),
                    'تاريخ الدخول': record.entry_date.strftime("%Y-%m-%d"),
                    'تاريخ الخروج': record.exit_date.strftime("%Y-%m-%d") if record.exit_date else "ما زالت في الورشة",
                    'حالة الإصلاح': status_map.get(record.repair_status, record.repair_status),
                    'التكلفة (ريال)': record.cost,
                    'اسم الورشة': record.workshop_name or "غير محدد",
                    'الفني المسؤول': record.technician_name or "غير محدد",
                    'رابط التسليم': record.delivery_link or "",
                    'رابط الاستلام': record.reception_link or "",
                    'الوصف': record.description or "",
                    'ملاحظات': record.notes or ""
                })
            
            df_workshop = pd.DataFrame(workshop_data)
            df_workshop.to_excel(writer, sheet_name='سجلات الورشة', index=False)
            
            # تعديل عرض الأعمدة
            worksheet = writer.sheets['سجلات الورشة']
            worksheet.set_column('A:A', 15)
            worksheet.set_column('B:C', 15)
            worksheet.set_column('D:D', 15)
            worksheet.set_column('E:E', 15)
            worksheet.set_column('F:G', 20)
            worksheet.set_column('H:I', 25)
            worksheet.set_column('J:K', 30)
            
            # تنسيق الخلايا
            for col_num, col in enumerate(df_workshop.columns):
                worksheet.write(0, col_num, col, header_format)
            
            for row_num in range(len(df_workshop)):
                for col_num in range(len(df_workshop.columns)):
                    value = df_workshop.iloc[row_num, col_num]
                    # إذا كان الحقل يحتوي على رابط، نقوم بإضافة تنسيق الرابط
                    if col_num in [7, 8] and isinstance(value, str) and value.startswith('http'):
                        url_format = workbook.add_format({
                            'font_color': 'blue',
                            'underline': 1,
                            'align': 'right',
                            'valign': 'vcenter',
                            'border': 1
                        })
                        worksheet.write_url(row_num + 1, col_num, value, url_format, string='اضغط هنا للفتح')
                    else:
                        worksheet.write(row_num + 1, col_num, value, cell_format)
            
            # إضافة صف إجمالي
            total_row = len(df_workshop) + 2
            worksheet.write(total_row, 0, 'الإجمالي', header_format)
            worksheet.write_formula(total_row, 4, f'=SUM(E2:E{len(df_workshop) + 1})', header_format)
            
            # دمج الخلايا للإجمالي
            worksheet.merge_range(total_row, 0, total_row, 3, 'الإجمالي', header_format)
        else:
            # إنشاء ورقة فارغة إذا لم تكن هناك سجلات
            worksheet = workbook.add_worksheet('سجلات الورشة')
            worksheet.write(0, 0, "لا توجد سجلات ورشة لهذه السيارة", header_format)
    
    buffer.seek(0)
    return buffer


def export_all_vehicles_to_excel():
    """
    تصدير جميع بيانات المركبات إلى ملف Excel شامل
    
    Returns:
        Flask Response: استجابة Flask مع ملف Excel
    """
    from flask import make_response, flash, redirect, url_for
    from models import Vehicle, VehicleWorkshop
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    try:
        # الحصول على جميع المركبات
        vehicles = Vehicle.query.order_by(Vehicle.plate_number).all()
        
        if not vehicles:
            flash("لا توجد مركبات للتصدير!", "warning")
            return redirect(url_for("vehicles.index"))
        
        # إنشاء مصنف Excel جديد
        wb = openpyxl.Workbook()
        
        # إزالة الورقة الافتراضية
        wb.remove(wb.active)
        
        # إنشاء ورقة المركبات الأساسية
        vehicles_ws = wb.create_sheet("المركبات")
        
        # إعداد التنسيق
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # عناوين الأعمدة للمركبات
        vehicle_headers = [
            "رقم اللوحة", "الماركة", "الموديل", "سنة الصنع", "اللون",
            "اسم السائق", "الحالة", "تاريخ انتهاء الفحص الدوري",
            "تاريخ انتهاء الاستمارة", "تاريخ انتهاء التفويض",
            "ملاحظات", "تاريخ الإضافة"
        ]
        
        # إضافة العناوين
        for col, header in enumerate(vehicle_headers, 1):
            cell = vehicles_ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # إضافة بيانات المركبات
        for row, vehicle in enumerate(vehicles, 2):
            vehicles_ws.cell(row=row, column=1, value=vehicle.plate_number or "")
            vehicles_ws.cell(row=row, column=2, value=vehicle.make or "")
            vehicles_ws.cell(row=row, column=3, value=vehicle.model or "")
            vehicles_ws.cell(row=row, column=4, value=vehicle.year or "")
            vehicles_ws.cell(row=row, column=5, value=vehicle.color or "")
            vehicles_ws.cell(row=row, column=6, value=vehicle.driver_name or "")
            
            # ترجمة حالة المركبة
            status_map = {
                "available": "متاحة",
                "rented": "مؤجرة",
                "in_use": "في الاستخدام",
                "maintenance": "في الصيانة",
                "in_workshop": "في الورشة",
                "in_project": "في المشروع",
                "accident": "حادث",
                "sold": "مباعة"
            }
            vehicles_ws.cell(row=row, column=7, value=status_map.get(vehicle.status, vehicle.status or ""))
            
            # تواريخ انتهاء الوثائق
            vehicles_ws.cell(row=row, column=8, value=vehicle.inspection_expiry_date.strftime("%Y-%m-%d") if vehicle.inspection_expiry_date else "")
            vehicles_ws.cell(row=row, column=9, value=vehicle.registration_expiry_date.strftime("%Y-%m-%d") if vehicle.registration_expiry_date else "")
            vehicles_ws.cell(row=row, column=10, value=vehicle.authorization_expiry_date.strftime("%Y-%m-%d") if vehicle.authorization_expiry_date else "")
            vehicles_ws.cell(row=row, column=11, value=vehicle.notes or "")
            vehicles_ws.cell(row=row, column=12, value=vehicle.created_at.strftime("%Y-%m-%d") if vehicle.created_at else "")
        
        # تعديل عرض الأعمدة
        for col in range(1, len(vehicle_headers) + 1):
            vehicles_ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # حفظ الملف في الذاكرة
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # إنشاء الاستجابة
        response = make_response(buffer.getvalue())
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        response.headers["Content-Disposition"] = f"attachment; filename=تصدير_جميع_المركبات_{timestamp}.xlsx"
        
        return response
        
    except Exception as e:
        flash(f"خطأ في تصدير البيانات: {str(e)}", "danger")
        return redirect(url_for("vehicles.index"))
